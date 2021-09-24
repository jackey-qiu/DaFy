import sys,os, qdarkstyle
import traceback
from io import StringIO
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox, QDialog
from PyQt5 import uic, QtWidgets
import PyQt5
import random
import numpy as np
import pandas as pd
import types,copy
import matplotlib.pyplot as plt
try:
    from . import locate_path
except:
    import locate_path
script_path = locate_path.module_path_locator()
DaFy_path = os.path.dirname(os.path.dirname(script_path))
sys.path.append(DaFy_path)
sys.path.append(os.path.join(DaFy_path,'dump_files'))
sys.path.append(os.path.join(DaFy_path,'EnginePool'))
sys.path.append(os.path.join(DaFy_path,'FilterPool'))
sys.path.append(os.path.join(DaFy_path,'util'))
from models.structure_tools.sxrd_dafy import lattice
from UtilityFunctions import locate_tag, replace_block, q_correction_for_one_Bragg_peak, q_correction_for_one_rod, fit_q_correction
from UtilityFunctions import apply_modification_of_code_block as script_block_modifier
from models.structure_tools.sxrd_dafy import AtomGroup
from models.utils import UserVars
import diffev
from diffev import fit_model_NLLS
from fom_funcs import *
import parameters
import data_superrod as data
import model
import solvergui
import time
import datetime
import matplotlib
# matplotlib.use("TkAgg")
matplotlib.rc('image', cmap='prism')
os.environ["QT_MAC_WANTS_LAYER"] = "1"
#import _tkinter
import pyqtgraph as pg
import pyqtgraph.exporters
from PyQt5 import QtCore
from PyQt5.QtWidgets import QCheckBox, QRadioButton, QTableWidgetItem, QHeaderView, QAbstractItemView, QInputDialog
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QTransform, QFont, QBrush, QColor, QIcon
from pyqtgraph.Qt import QtGui
from threading import Thread
import syntax_pars
from models.structure_tools import sorbate_tool
from models.structure_tools import sorbate_tool_beta
import logging
from superrod_ui import Ui_MainWindow

class PandasModel(QtCore.QAbstractTableModel):
    """
    Class to populate a table view with a pandas dataframe
    """
    def __init__(self, data, tableviewer, main_gui, parent=None):
        QtCore.QAbstractTableModel.__init__(self, parent)
        self._data = data
        self.tableviewer = tableviewer
        self.main_gui = main_gui

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parent=None):
        return self._data.shape[1]

    def data(self, index, role):
        if index.isValid():
            if role in [QtCore.Qt.DisplayRole, QtCore.Qt.EditRole]:
                return str(self._data.iloc[index.row(), index.column()])
            if role == QtCore.Qt.BackgroundRole and index.row()%2 == 0:
                return QtGui.QColor('DeepSkyBlue')
                # return QtGui.QColor('green')
            if role == QtCore.Qt.BackgroundRole and index.row()%2 == 1:
                return QtGui.QColor('aqua')
                # return QtGui.QColor('lightGreen')
            if role == QtCore.Qt.ForegroundRole and index.row()%2 == 1:
                return QtGui.QColor('black')
            '''
            if role == QtCore.Qt.CheckStateRole and index.column()==0:
                if self._data.iloc[index.row(),index.column()]:
                    return QtCore.Qt.Checked
                else:
                    return QtCore.Qt.Unchecked
            '''
        return None

    def setData(self, index, value, role):
        if not index.isValid():
            return False
        '''
        if role == QtCore.Qt.CheckStateRole and index.column() == 0:
            if value == QtCore.Qt.Checked:
                self._data.iloc[index.row(),index.column()] = True
            else:
                self._data.iloc[index.row(),index.column()] = False
        else:
        '''
        if str(value)!='':
            self._data.iloc[index.row(),index.column()] = str(value)
        #if self._data.columns.tolist()[index.column()] in ['select','archive_data','user_label','read_level']:
        #    self.main_gui.update_meta_info_paper(paper_id = self._data['paper_id'][index.row()])
        self.dataChanged.emit(index, index)
        self.layoutAboutToBeChanged.emit()
        self.dataChanged.emit(self.createIndex(0, 0), self.createIndex(self.rowCount(0), self.columnCount(0)))
        self.layoutChanged.emit()
        self.tableviewer.resizeColumnsToContents() 
        return True

    def update_view(self):
        self.layoutAboutToBeChanged.emit()
        self.dataChanged.emit(self.createIndex(0, 0), self.createIndex(self.rowCount(0), self.columnCount(0)))
        self.layoutChanged.emit()

    def headerData(self, rowcol, orientation, role):
        if orientation == QtCore.Qt.Horizontal and role == QtCore.Qt.DisplayRole:
            return self._data.columns[rowcol]         
        if orientation == QtCore.Qt.Vertical and role == QtCore.Qt.DisplayRole:
            return self._data.index[rowcol]         
        return None

    def flags(self, index):
        if not index.isValid():
           return QtCore.Qt.NoItemFlags
        else:
            return (QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEditable)

    def sort(self, Ncol, order):
        """Sort table by given column number."""
        self.layoutAboutToBeChanged.emit()
        self._data = self._data.sort_values(self._data.columns.tolist()[Ncol],
                                        ascending=order == QtCore.Qt.AscendingOrder, ignore_index = True)
        self.dataChanged.emit(self.createIndex(0, 0), self.createIndex(self.rowCount(0), self.columnCount(0)))
        self.layoutChanged.emit()

class DummydataGeneraterDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        # Load the dialog's GUI
        uic.loadUi(os.path.join(script_path,"generate_dummy_data_gui.ui"), self)
        self.setWindowTitle('Dummy data Generator in an easy way')

        self.pushButton_extract_ctr.clicked.connect(self.extract_ctr_set_table)
        self.pushButton_extract_raxs.clicked.connect(self.extract_raxs_set_table)
        self.pushButton_generate.clicked.connect(self.generate_dummy_data)

    def extract_ctr_set_table(self):
        rows = self.spinBox_ctr.value()
        data_ = {'h':[0]*rows, 'k':[0]*rows, 'l_min':[0]*rows, 'l_max':[5]*rows, 'delta_l': [0.1]*rows, 'Bragg_Ls':[str([2,4,6])]*rows}
        self.pandas_model_ctr = PandasModel(data = pd.DataFrame(data_), tableviewer = self.tableView_ctr, main_gui = self.parent)
        self.tableView_ctr.setModel(self.pandas_model_ctr)
        self.tableView_ctr.resizeColumnsToContents()
        self.tableView_ctr.setSelectionBehavior(PyQt5.QtWidgets.QAbstractItemView.SelectRows)

    def extract_raxs_set_table(self):
        rows = self.spinBox_raxs.value()
        data_ = {'h':[0]*rows, 'k':[0]*rows, 'l':[0.3]*rows, 'E_min':[float(self.lineEdit_E_min.text())]*rows, 'E_max':[float(self.lineEdit_E_max.text())]*rows, 'delta_E': [1]*rows, 'E0':[float(self.lineEdit_E0.text())]*rows}
        self.pandas_model_raxs = PandasModel(data = pd.DataFrame(data_), tableviewer = self.tableView_raxs, main_gui = self.parent)
        self.tableView_raxs.setModel(self.pandas_model_raxs)
        self.tableView_raxs.resizeColumnsToContents()
        self.tableView_raxs.setSelectionBehavior(PyQt5.QtWidgets.QAbstractItemView.SelectRows)

    @staticmethod
    def _generate_dummy_data(data,file_path):
        np.savetxt(file_path,data)

    def generate_dummy_data(self):
        from functools import partial
        data_ctr = self._extract_hkl_all()
        data_raxs = self._extract_hkl_all_raxs()
        if len(data_raxs)!=0:
            kwargs = {'ctr':data_ctr,'raxs':data_raxs,'func':partial(self._generate_dummy_data, file_path = self.lineEdit_data_path.text())}
        else:
            kwargs = {'ctr':data_ctr,'func':partial(self._generate_dummy_data, file_path = self.lineEdit_data_path.text())}
        self.parent.model.script_module.Sim(self.parent.model.script_module.data, kwargs = kwargs)

    def _extract_hkl_one_rod(self, h, k, l_min, l_max, Bragg_ls, delta_l):
        ls = np.linspace(l_min, l_max, int((l_max-l_min)/delta_l))
        for each in Bragg_ls:
            ls = ls[abs(ls - each)>delta_l]
        hs = np.array([h]*len(ls))
        ks = np.array([k]*len(ls))
        return np.array([hs,ks,ls]).T

    def _extract_hkl_one_raxs(self, h, k, l, E_min, E_max, delta_E, E0):
        Es_left = list(np.linspace(E_min, E0-20, int((E0-20-E_min)/10)))
        Es_edge = list(np.linspace(E0-20, E0+20, int(40/delta_E)))
        Es_right = list(np.linspace(E0+20, E_max, int((E_max-E0-20)/10)))
        Es = Es_left + Es_edge + Es_right
        #ls = [l]*len(Es)
        hs = [h]*len(Es)
        ks = [k]*len(Es)
        ls = [l]*len(Es)
        #dls = [2]*len(Es)
        #Bls = [2]*len(Es)
        #fs_ = self.parent.model.script_module.sample.calc_f_all_RAXS(np.array(hs), np.array(ks), np.array(ys), np.array(Es))
        #fs = abs(fs_*fs_)
        #errs = fs*0.005
        return np.array([hs,ks,ls, Es]).T

    def _extract_hkl_all(self):
        rows = len(self.pandas_model_ctr._data.index)
        data_temp = np.zeros((1,3))[0:0]
        for i in range(rows):
            h, k, l_min, l_max, delta_l, Bragg_ls = self.pandas_model_ctr._data.iloc[i,:].tolist()
            Bragg_ls = eval(Bragg_ls)
            result = self._extract_hkl_one_rod(int(h),int(k),float(l_min),float(l_max),Bragg_ls, float(delta_l))
            data_temp = np.vstack((data_temp,result))
        return data_temp

    def _extract_hkl_all_raxs(self):
        rows = len(self.pandas_model_raxs._data.index)
        data_temp = np.zeros((1,4))[0:0]
        for i in range(rows):
            h, k, l, E_min, E_max, delta_E, E0 = self.pandas_model_raxs._data.iloc[i,:].tolist()
            result = self._extract_hkl_one_raxs(int(h),int(k),float(l),float(E_min),float(E_max),float(delta_E), float(E0))
            data_temp = np.vstack((data_temp,result))
        return data_temp

class ScriptGeneraterDialog(QDialog):
    def __init__(self, parent=None):
        # print(sorbate_tool_beta.STRUCTURE_MOTIFS)
        super().__init__(parent)
        self.parent = parent
        
        # Load the dialog's GUI
        uic.loadUi(os.path.join(script_path,"ctr_model_creator.ui"), self)
        self.setWindowTitle('Script Generator in an easy way')
        self.plainTextEdit_script.setStyleSheet("""QPlainTextEdit{
                        font-family:'Consolas';
                        font-size:14pt;
                        color: #ccc;
                        background-color: #2b2b2b;}""")
        self.plainTextEdit_script.setTabStopWidth(self.plainTextEdit_script.fontMetrics().width(' ')*4)
        #set combobox text items
        self.comboBox_motif_types.addItems(list(sorbate_tool_beta.STRUCTURE_MOTIFS.keys()))

        self.comboBox_predefined_symmetry.addItems(list(sorbate_tool_beta.SURFACE_SYMS.keys()))
        self.comboBox_predefined_symmetry.currentTextChanged.connect(self.reset_sym_info)
        self.pushButton_add_symmetry.clicked.connect(self.append_sym_info)
        self.pushButton_add_all.clicked.connect(self.append_all_sym)

        self.pushButton_extract_surface.clicked.connect(self.extract_surface_slabs)
        self.pushButton_generate_script_surface.clicked.connect(self.generate_script_surface_slabs_and_surface_atm_group)

        self.comboBox_predefined_subMotifs.clear()
        self.comboBox_predefined_subMotifs.addItems(sorbate_tool_beta.STRUCTURE_MOTIFS[self.comboBox_motif_types.currentText()])
        self.comboBox_motif_types.currentTextChanged.connect(self.reset_combo_motif)
        self.pushButton_make_setting_table.clicked.connect(self.setup_sorbate_setting_table)
        self.pushButton_apply_setting.clicked.connect(self.apply_settings_for_one_sorbate)
        self.pushButton_generate_script_sorbate.clicked.connect(self.generate_script_snippet_sorbate)

        self.pushButton_generate_full_script.clicked.connect(self.generate_full_script)
        self.pushButton_transfer_script.clicked.connect(self.transfer_script)

        self.pushButton_draw_structure.clicked.connect(self.show_3d_structure)
        self.pushButton_pan.clicked.connect(self.pan_view)

        self.pushButton_hematite.clicked.connect(self.load_hematite)
        self.pushButton_mica.clicked.connect(self.load_mica)
        self.pushButton_cu.clicked.connect(self.load_cu)

        self.script_lines_sorbate = {}
        self.script_lines_update_sorbate = {'update_sorbate':[]}
        self.script_container = {}
        self.lineEdit_bulk.setText(os.path.join(DaFy_path,'util','batchfile','Cu100','Cu100_bulk.str'))
        self.lineEdit_folder_suface.setText(os.path.join(DaFy_path,'util','batchfile','Cu100'))
        self.lineEdit_template_script.setText(os.path.join(script_path,'standard_scripts','template_script.py'))

    def load_hematite(self):
        self.lineEdit_bulk.setText(os.path.join(DaFy_path,'util','batchfile','hematite_rcut','bulk.str'))
        self.lineEdit_folder_suface.setText(os.path.join(DaFy_path,'util','batchfile','hematite_rcut'))
        self.lineEdit_files_surface.setText('half_layer2.str')
        self.lineEdit_lattice.setText(str([5.038,5.434,7.3707,90,90,90]))
        self.lineEdit_surface_offset.setText(str({'delta1':0.,'delta2':0.1391}))
        self.lineEdit_template_script.setText(os.path.join(DaFy_path,'projects','superrod','standard_scripts','template_script.py'))
        self.comboBox_T_factor.setCurrentText('B')

    def load_cu(self):
        self.lineEdit_bulk.setText(os.path.join(DaFy_path,'util','batchfile','Cu100','Cu100_bulk.str'))
        self.lineEdit_folder_suface.setText(os.path.join(DaFy_path,'util','batchfile','Cu100'))
        self.lineEdit_files_surface.setText('Cu100_surface_1.str')
        self.lineEdit_lattice.setText(str([3.615,3.615,3.615,90,90,90]))
        self.lineEdit_surface_offset.setText(str({'delta1':0.,'delta2':0.}))
        self.lineEdit_template_script.setText(os.path.join(DaFy_path,'projects','superrod','standard_scripts','template_script.py'))
        self.comboBox_T_factor.setCurrentText('u')

    def load_mica(self):
        self.lineEdit_bulk.setText(os.path.join(DaFy_path,'util','batchfile','Muscovite001','muscovite_001_bulk_u_corrected_new.str'))
        self.lineEdit_folder_suface.setText(os.path.join(DaFy_path,'util','batchfile','Muscovite001'))
        self.lineEdit_files_surface.setText('muscovite_001_surface_AlSi_u_corrected_new_1.str')
        self.lineEdit_lattice.setText(str([5.1988,9.0266,20.04156,90,95.782,90]))
        self.lineEdit_surface_offset.setText(str({'delta1':0.,'delta2':0.}))
        self.lineEdit_template_script.setText(os.path.join(DaFy_path,'projects','superrod','standard_scripts','template_script.py'))
        self.comboBox_T_factor.setCurrentText('u')

    def show_3d_structure(self):
        self.lattice = lattice(*eval(self.lineEdit_lattice.text()))
        self.widget_structure.T = self.lattice.RealTM
        self.widget_structure.T_INV = self.lattice.RealTMInv
        self.widget_structure.show_bond_length = True
        self.widget_structure.clear()
        self.widget_structure.opts['distance'] = 2000
        self.widget_structure.opts['fov'] = 1
        self.widget_structure.abc = np.array(eval(self.lineEdit_lattice.text())[0:3])
        xyz_init = self.pandas_model_slab._data[self.pandas_model_slab._data['show']=='1'][['el','id','x','y','z']]
        translation_offsets = [np.array([0,0,0]),np.array([1,0,0]),np.array([-1,0,0]),np.array([0,1,0]),np.array([0,-1,0]),np.array([1,-1,0]),np.array([-1,1,0]),np.array([1,1,0]),np.array([-1,-1,0])]
        xyz_ = pd.DataFrame({'el':[],'id':[],'x':[],'y':[],'z':[]})
        xyz = []
        for i in range(len(xyz_init.index)):
            el, id, x, y, z = xyz_init.iloc[i].tolist()
            for tt in translation_offsets:
                i, j, k = tt
                tag = '_'
                if i!=0:
                    tag=tag+'{}x'.format(i)
                if j!=0:
                    tag=tag+'{}y'.format(j)
                if k!=0:
                    tag=tag+'{}z'.format(k)
                if tag == '_':
                    tag = ''
                _x, _y, _z = x+i, y+j, z+k
                _id = id+tag
                xyz_.loc[len(xyz_.index)] = [el, _id, _x, _y, _z]
                x_c, y_c, z_c = np.dot(self.widget_structure.T,np.array([_x,_y,_z]))
                xyz.append([el,_id, x_c, y_c, z_c])
        #xyz = list(zip(xyz_['el'].tolist(),xyz_['id'].tolist(),xyz_['x']*self.widget_structure.abc[0].tolist(),xyz_['y']*self.widget_structure.abc[1].tolist(),xyz_['z']*self.widget_structure.abc[2].tolist()))
        self.widget_structure.show_structure(xyz, show_id = True)

    def pan_view(self):
        value = int(self.spinBox_pan_pixel.text())
        self.widget_structure.pan(value*int(self.checkBox_x.isChecked()),value*int(self.checkBox_y.isChecked()),value*int(self.checkBox_z.isChecked()))

    def reset_sym_info(self):
        self.lineEdit_2d_rotation_matrix.setText(str(sorbate_tool_beta.SURFACE_SYMS[self.comboBox_predefined_symmetry.currentText()][0:2]))
        self.lineEdit_translation_offset.setText(str(sorbate_tool_beta.SURFACE_SYMS[self.comboBox_predefined_symmetry.currentText()][2]))

    def append_sym_info(self):
        text = self.textEdit_symmetries.toPlainText()
        if text == '':
            self.textEdit_symmetries.setPlainText(f"model.SymTrans({self.lineEdit_2d_rotation_matrix.text()},t={self.lineEdit_translation_offset.text()})")
        else:
            self.textEdit_symmetries.setPlainText(text+f"\nmodel.SymTrans({self.lineEdit_2d_rotation_matrix.text()},t={self.lineEdit_translation_offset.text()})")

    def append_all_sym(self):
        sym_symbols = [self.comboBox_predefined_symmetry.itemText(i) for i in range(self.comboBox_predefined_symmetry.count())]
        for each in sym_symbols:
            mt = str(sorbate_tool_beta.SURFACE_SYMS[each][0:2])
            tt = str(sorbate_tool_beta.SURFACE_SYMS[each][2])
            text = self.textEdit_symmetries.toPlainText()
            if text=='':
                self.textEdit_symmetries.setPlainText(f"model.SymTrans({mt},t={tt})")
            else:
                self.textEdit_symmetries.setPlainText(text+f"\nmodel.SymTrans({mt},t={tt})")

    def reset_combo_motif(self):
        self.comboBox_predefined_subMotifs.clear()
        self.comboBox_predefined_subMotifs.addItems(sorbate_tool_beta.STRUCTURE_MOTIFS[self.comboBox_motif_types.currentText()])

    def setup_sorbate_setting_table(self):
        module = getattr(sorbate_tool_beta,self.comboBox_motif_types.currentText())
        data_ = {}
        #if self.checkBox_use_predefined.isChecked():
        data_['sorbate'] = [str(self.spinBox_sorbate_index.value())]
        data_['motif'] = [self.comboBox_predefined_subMotifs.currentText()]
        data_.update(module.get_par_dict(self.comboBox_predefined_subMotifs.currentText()))
        '''
        else:
            data_['sorbate'] = [str(self.spinBox_sorbate_index.value())]
            data_['xyzu_oc_m'] = str([0.5, 0.5, 1.5, 0.1, 1, 1])
            data_['els'] = str(['O','C','C','O'])
            data_['flat_down_index'] = str([2])
            data_['anchor_index_list'] = str([1, None, 1, 2])
            data_['lat_pars'] = str([3.615, 3.615, 3.615, 90, 90, 90])
            data_['structure_pars_dict'] = str({'r':1.5, 'delta':0})
            data_['binding_mode'] = 'OS'
            data_['structure_index'] = str(self.spinBox_sorbate_index.value())
        '''
        self.pandas_model = PandasModel(data = pd.DataFrame(data_), tableviewer = self.tableView_sorbate_setting, main_gui = self.parent)
        self.tableView_sorbate_setting.setModel(self.pandas_model)
        self.tableView_sorbate_setting.resizeColumnsToContents()
        self.tableView_sorbate_setting.setSelectionBehavior(PyQt5.QtWidgets.QAbstractItemView.SelectRows)

    def apply_settings_for_one_sorbate(self):
        module = getattr(sorbate_tool_beta,self.comboBox_motif_types.currentText())
        # module = eval(f"sorbate_tool_beta.{self.comboBox_motif_types.currentText()}")
        '''
        if self.checkBox_use_predefined.isChecked():
            results = module.generate_script_from_setting_table(use_predefined_motif = True, predefined_motif = self.pandas_model._data.iloc[0]['motif'], structure_index = self.pandas_model._data.iloc[0]['sorbate'])
            self.script_lines_sorbate[self.pandas_model._data.iloc[0]['sorbate']]=results[0]
            self.script_lines_update_sorbate['update_sorbate'].append(results[1])
        else:
        '''
        kwargs = {each:self.pandas_model._data.iloc[0][each] for each in self.pandas_model._data.columns}
        results = module.generate_script_from_setting_table(use_predefined_motif = False, structure_index = self.pandas_model._data.iloc[0]['sorbate'], kwargs = kwargs)
        self.script_lines_sorbate[self.pandas_model._data.iloc[0]['sorbate']]=results[0]
        self.script_lines_update_sorbate['update_sorbate'].append(results[1])
        self.reset_sorbate_set()

    def reset_sorbate_set(self):
        self.lineEdit_sorbate_set.setText(','.join(['sorbate_{}'.format(each) for each in self.script_lines_sorbate.keys()]))

    def generate_script_snippet_sorbate(self):
        keys = sorted(list(self.script_lines_sorbate.keys()))
        scripts = '\n\n'.join([self.script_lines_sorbate[each] for each in keys])
        syms = ''
        for each in keys:
            syms= syms+ "sorbate_syms_{}=[{}]\n".format(each,','.join(self.textEdit_symmetries.toPlainText().rsplit('\n')))
        self.script_container['sorbateproperties'] = scripts+'\n'
        self.script_container['sorbatesym'] = syms
        self.script_container['update_sorbate'] = '\n'.join(list(set(self.script_lines_update_sorbate['update_sorbate'])))+'\n'
        if 'slabnumber' not in self.script_container:
            self.script_container['slabnumber'] = {'num_sorbate_slabs':str(len(keys))}
        else:
            self.script_container['slabnumber'].update({'num_sorbate_slabs':str(len(keys))})

        self.plainTextEdit_script.setPlainText(scripts+syms)

    def extract_surface_slabs(self):
        files = [os.path.join(self.lineEdit_folder_suface.text(),each) for each in self.lineEdit_files_surface.text().rsplit()]
        def _make_df(file, slab_index):
            df = pd.read_csv(file, comment = '#', names = ['id','el','x','y','z','u','occ','m'])
            df['slab'] = slab_index
            df['show'] = str(0)
            df['sym_matrix']=str([1,0,0,0,1,0,0,0,1])
            df['gp_tag'] = 'NaN'
            return df
        dfs = []
        for i in range(len(files)):
            dfs.append(_make_df(files[i],i+1))
        df = pd.concat(dfs, ignore_index = True)
        df.sort_values(by = ['slab','z','id'], ascending = False, inplace = True, ignore_index=True)
        self.pandas_model_slab = PandasModel(data = pd.DataFrame(df), tableviewer = self.tableView_surface_slabs, main_gui = self.parent)
        self.tableView_surface_slabs.setModel(self.pandas_model_slab)
        self.tableView_surface_slabs.resizeColumnsToContents()
        self.tableView_surface_slabs.setSelectionBehavior(PyQt5.QtWidgets.QAbstractItemView.SelectRows)

    def generate_script_surface_slabs(self):
        scripts = []
        files = [os.path.join(self.lineEdit_folder_suface.text(),each) for each in self.lineEdit_files_surface.text().rsplit()]
        for i in range(len(files)):
            scripts.append("surface_{} = model.Slab(T_factor = '{}')".format(i+1,self.comboBox_T_factor.currentText()))
            scripts.append("tool_box.add_atom_in_slab(surface_{}, '{}')".format(i+1, files[i]))
        # scripts.append('\n')
        self.script_container['surfaceslab'] = '\n'.join(scripts) + '\n'
        #self.script_container['slabnumber'] = {'num_surface_slabs':str(len(files))}
        if 'slabnumber' not in self.script_container:
            self.script_container['slabnumber'] = {'num_surface_slabs':str(len(files))}
        else:
            self.script_container['slabnumber'].update({'num_surface_slabs':str(len(files))})
        return self.script_container['surfaceslab']

    def generate_script_surface_atm_group(self):
        scripts = []
        atm_gps_df = self.pandas_model_slab._data[self.pandas_model_slab._data['gp_tag']!='NaN']
        gp_tags = sorted(list(set(atm_gps_df['gp_tag'].tolist())))
        for each in gp_tags:
            scripts.append("{} = model.AtomGroup(instance_name = '{}')".format(each, each))
            temp = atm_gps_df[atm_gps_df['gp_tag']==each]
            for i in range(temp.shape[0]):
                scripts.append("{}.add_atom({},'{}',matrix={})".format(each,'surface_'+str(temp.iloc[i]['slab']),temp.iloc[i]['id'],str(temp.iloc[i]['sym_matrix'])))
        # scripts.append('\n')
        self.script_container['atmgroup'] = '\n'.join(scripts)
        return self.script_container['atmgroup']

    def generate_script_surface_slabs_and_surface_atm_group(self):
        self.plainTextEdit_script.setPlainText(self.generate_script_surface_slabs()+'\n\n'+self.generate_script_surface_atm_group())

    def generate_script_bulk(self):
        self.script_container['sample'] = {'surface_parms':self.lineEdit_surface_offset.text()}
        if os.path.isfile(self.lineEdit_bulk.text()):
            self.script_container['bulk'] = "bulk = model.Slab(T_factor = '{}')\ntool_box.add_atom_in_slab(bulk,'{}')\n".format(self.comboBox_T_factor.currentText(),self.lineEdit_bulk.text())

    def generate_script_raxs(self):
        self.script_container['raxs']  = "RAXS_EL = '{}'\nRAXS_FIT_MODE = '{}'\nNUMBER_SPECTRA = {}\nE0 = {}\nF1F2_FILE = '{}'\n".format(self.lineEdit_res_el.text(),
                                                                                                               self.comboBox_mode.currentText(),
                                                                                                               str(self.spinBox_num_raxs.value()),
                                                                                                               self.lineEdit_e0.text(),
                                                                                                               self.lineEdit_f1f2.text())


    def generate_full_script(self):
        with open(self.lineEdit_template_script.text(),'r') as f:
            lines = f.readlines()
            if len(lines)== 0:
                print('There are 0 lines in the file!')
                return
            #bulk file
            self.generate_script_bulk()
            #raxs setting
            self.generate_script_raxs()
            #lattice parameters
            self.script_container['unitcell']={'lat_pars':eval(self.lineEdit_lattice.text())}
            #energy
            self.script_container['wavelength']={'wal':round(12.398/eval(self.lineEdit_E.text()),4)}

            #surface slabs
            self.generate_script_surface_slabs()
            #surface atm groups
            self.generate_script_surface_atm_group()
            #sorbate and symmetry
            self.generate_script_snippet_sorbate()
            ##Now let us modify the script
            for key in self.script_container:
                if type(self.script_container[key])==type({}):
                    lines = script_block_modifier(lines, key, list(self.script_container[key].keys()), list(self.script_container[key].values()))
                else:
                    lines = replace_block(lines, key, self.script_container[key])
            self.plainTextEdit_script.setPlainText(''.join(lines))

    def transfer_script(self):
        self.parent.plainTextEdit_script.setPlainText(self.plainTextEdit_script.toPlainText())

#redirect the error stream to qt widget
class QTextEditLogger(logging.Handler):
    def __init__(self, textbrowser_widget):
        super().__init__()
        self.textBrowser_error_msg = textbrowser_widget
        # self.widget.setReadOnly(True)

    def emit(self, record):
        error_msg = self.format(record)
        separator = '-' * 80
        notice = \
        """An unhandled exception occurred. Please report the problem\n"""\
        """using the error reporting dialog or via email to <%s>.\n"""%\
        ("crqiu2@gmail.com")
        self.textBrowser_error_msg.clear()
        cursor = self.textBrowser_error_msg.textCursor()
        cursor.insertHtml('''<p><span style="color: red;">{} <br></span>'''.format(" "))
        self.textBrowser_error_msg.setText(notice + '\n' +separator+'\n'+error_msg)

class ScanPar(QtCore.QObject):
    def __init__(self,model):
        super(ScanPar, self).__init__()
        self.model = model
        self.running = False
        self.sign = 1
        self.row = 0
        self.steps = 10

    def run(self):
        self.running = True
        while True:
            # print('Running!')
            if self.running:
                self.rock_par()
                time.sleep(1.5)
            else:
                break

    def rock_par(self):
        par_set = self.model.parameters.data[self.row]
        par_min, par_max = par_set[-4], par_set[-3]
        #steps = int(self.spinBox_steps.value())
        steps = self.steps
        old_value = self.model.parameters.get_value(self.row, 1)
        new_value = (par_max - par_min)/steps*self.sign + old_value
        if (new_value>par_max) or (new_value<par_min):
            self.model.parameters.set_value(self.row, 1, new_value - 2*self.sign*(par_max - par_min)/steps)
            self.sign = -self.sign
        else:
            self.model.parameters.set_value(self.row, 1, new_value)

    def stop(self):
        self.running = False

class RunFit(QtCore.QObject):
    """
    RunFit class to interface GUI to operate fit-engine, which is ran on a different thread
    ...
    Attributes
    ----------
    updateplot : pyqtSignal be emitted to be received by main GUI thread during fit
    solver: api for model fit using differential evolution algorithm

    Methods
    ----------
    run: start the fit
    stop: stop the fit

    """
    updateplot = QtCore.pyqtSignal(str,object,bool)
    fitended = QtCore.pyqtSignal(str)
    def __init__(self,solver):
        super(RunFit, self).__init__()
        self.solver = solver
        self.running = False

    def run(self):
        self.running = True
        self.solver.optimizer.stop = False
        self.solver.StartFit(self.updateplot,self.fitended)

    def stop(self):
        self.running = False
        self.solver.optimizer.stop = True

class RunBatch(QtCore.QObject):
    """
    RunFit class to interface GUI to operate fit-engine, which is ran on a different thread
    ...
    Attributes
    ----------
    updateplot : pyqtSignal be emitted to be received by main GUI thread during fit
    solver: api for model fit using differential evolution algorithm
    multiple_files_hooker: 
        True if you want to run sequentially the rod files in the listWidget
        False if you want to rolling the fit on one rod files (used in fitting many RAXR spectrum)

    Methods
    ----------
    run: start the fit
    stop: stop the fit

    """
    updateplot = QtCore.pyqtSignal(str,object,bool)
    fitended = QtCore.pyqtSignal(str)
    def __init__(self,solver):
        super(RunBatch, self).__init__()
        self.solver = solver
        self.running = False
        self.multiple_files_hooker = False

    def run(self):
        self.running = True
        self.solver.optimizer.stop = False
        self.solver.StartFit(self.updateplot,self.fitended)

    def set_hooker(self,hooker):
        self.multiple_files_hooker = hooker

    def stop(self):
        self.running = False
        self.solver.optimizer.stop = True

class MyMainWindow(QMainWindow):
    """
    GUI class for this app
    ....
    Attributes (selected)
    -----------
    <<widgets>>
    tableWidget_data: QTableWidget holding a list of datasets
    tableWidget_data_view: QTableWidget displaying each dataset
    widget_solver:pyqtgraph.parameter_tree_widget where you define
                  intrinsic parameters for undertaking DE optimization
    tableWidget_pars: QTableWidget displaying fit parameters
    widget_data: pyqtgraph.GraphicsLayoutWidget showing figures of
                 each ctr profile (data, fit, ideal and errorbars)
    widget_fom: pyqtgraph.GraphicsLayoutWidget showing evolution of
                figure of merit during fit
    widget_pars:pyqtgraph.GraphicsLayoutWidget showing best fit of
                each parameter at current generation and the search
                range in bar chart at this moment. longer bar means
                more aggressive searching during fit. If the bars 
                converge to one narrow line, fit quality cannot improved
                anymore. That means the fit is finished.
    widget_edp: GLViewWidget showing the 3d molecular structure of the
                current best fit model.
    widget_msv_top: GLViewWidget showing the top view of 3d molecular
                structure of the current best fit model. Only one sorbate
                and one layer of surface atoms are shown for clarity.
    plainTextEdit_script: QCodeEditor widget showing the model script
    widget_terminal:TerminalWidget, where you can do whatever you can in
                a normal python terminal. Three variables are loaded in 
                the namespace of the terminal:
                1) win: GUI main frame
                2) model: model that bridget script_module, pars and Fit engine
                you can explore the variables defined in your model script
                using model.script_module.vars (if vars is defined in script)
    <<others>>
    run_fit: Run_Fit instance to be launched to start/stop a fit. Refer to
             Run_Fit.solver to learn more about implementation of multi-processing
             programe method.
    model: model instance to coordinate script name space, dataset instance and par
           instance
    f_ideal: a list holding the structure factor values for unrelaxed structure
    data_profile: list of handles to plot ctr profiles including data and fit reuslts

    Methods (selected)
    -----------
    """
    def __init__(self, parent = None):
        super(MyMainWindow, self).__init__(parent)
        # self.setupUi(self)
        #pyqtgraph preference setting
        pg.setConfigOptions(imageAxisOrder='row-major', background = (50,50,100))
        pg.mkQApp()
        #load GUI ui file made by qt designer
        uic.loadUi(os.path.join(DaFy_path,'projects','SuperRod','superrod_gui.ui'),self)
        self.widget_terminal.update_name_space("win",self)
        self.setWindowTitle('Data analysis factory: CTR data modeling')
        icon = QIcon(os.path.join(script_path,"icons","DAFY.png"))
        self.setWindowIcon(icon)

        #set redirection of error message to embeted text browser widget
        logTextBox = QTextEditLogger(self.textBrowser_error_msg)
        # You can format what is printed to text box
        logTextBox.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logging.getLogger().addHandler(logTextBox)
        # You can control the logging level
        logging.getLogger().setLevel(logging.DEBUG)

        # self.comboBox_all_motif.insertItems(0, sorbate_tool.ALL_MOTIF_COLLECTION)
        #self.stop = False
        self.show_checkBox_list = []
        self.domain_tag = 1
        #structure factor for ideal structure
        self.f_ideal=[]
        self.data_profiles = []
        self.model = model.Model()
        self.nlls_fit = fit_model_NLLS(self.model)
        
        self.timer_nlls = QTimer(self)
        self.timer_nlls.timeout.connect(self.update_status_nlls)
        #init run_fit
        self.run_fit = RunFit(solvergui.SolverController(self.model))
        self.fit_thread = QtCore.QThread()
        self.run_fit.moveToThread(self.fit_thread)#move run_fit to a different thread
        #signal-slot connection
        self.run_fit.updateplot.connect(self.update_par_during_fit)
        self.run_fit.updateplot.connect(self.update_status)
        self.run_fit.fitended.connect(self.stop_model_slot)
        self.fit_thread.started.connect(self.run_fit.run)

        #init run_batch
        self.run_batch = RunBatch(solvergui.SolverController(self.model))
        self.batch_thread = QtCore.QThread()
        self.run_batch.moveToThread(self.batch_thread)
        #signal-slot connection
        self.run_batch.updateplot.connect(self.update_par_during_fit)
        self.run_batch.updateplot.connect(self.update_status_batch)
        self.run_batch.fitended.connect(self.stop_model_batch)
        # self.run_batch.fitended.connect(lambda:self.timer_update_structure.stop())
        self.batch_thread.started.connect(self.run_batch.run)
        # self.batch_thread.started.connect(lambda:self.timer_update_structure.start(2000))

        self.scan_par = ScanPar(self.model)
        self.scan_par_thread = QtCore.QThread()
        self.scan_par.moveToThread(self.scan_par_thread)
        self.scan_par_thread.started.connect(self.scan_par.run)
        self.pushButton_scan.clicked.connect(self.start_scan_par_thread)
        self.pushButton_stop_scan.clicked.connect(self.stop_scan_par_thread)
        self.timer_scan_par = QTimer(self)
        self.timer_scan_par.timeout.connect(self.update_structure_during_scan_par)
        #tool bar buttons to operate model
        self.actionNew.triggered.connect(self.init_new_model)
        self.actionOpen.triggered.connect(self.open_model)
        self.actionSaveas.triggered.connect(self.save_model_as)
        self.actionSave.triggered.connect(self.save_model)
        self.actionSimulate.triggered.connect(lambda:self.simulate_model(compile = True))
        # self.actionCompile.triggered.connect(lambda:self.simulate_model(compile = False))
        self.actionRun.triggered.connect(self.run_model)
        self.actionStop.triggered.connect(self.stop_model)
        self.actionCalculate.triggered.connect(self.calculate_error_bars)
        self.actionRun_batch_script.triggered.connect(self.run_model_batch)
        self.actionStopBatch.triggered.connect(self.terminate_model_batch)

        #menu items
        self.actionOpen_model.triggered.connect(self.open_model)
        self.actionSave_model.triggered.connect(self.save_model_as)
        self.actionSimulate_2.triggered.connect(lambda:self.simulate_model(compile = True))
        self.actionStart_fit.triggered.connect(self.run_model)
        self.actionNLLS.triggered.connect(self.start_nlls)
        self.actionStop_fit.triggered.connect(self.stop_model)
        self.actionSave_table.triggered.connect(self.save_par)
        self.actionSave_script.triggered.connect(self.save_script)
        self.actionSave_data.triggered.connect(self.save_data)
        self.actionData.changed.connect(self.toggle_data_panel)
        self.actionPlot.changed.connect(self.toggle_plot_panel)
        self.actionScript.changed.connect(self.toggle_script_panel)

        self.pushButton_generate_script.clicked.connect(self.generate_script_dialog)

        #pushbuttons for model file navigator 
        self.pushButton_load_files.clicked.connect(self.load_rod_files)
        self.pushButton_clear_selected_files.clicked.connect(self.remove_selected_rod_files)
        self.listWidget_rod_files.itemDoubleClicked.connect(self.open_model_selected_in_listWidget)
        self.pushButton_open_selected_rod_file.clicked.connect(self.open_model_selected_in_listWidget)
        self.pushButton_hook_to_batch.clicked.connect(self.hook_to_batch)
        self.pushButton_purge_from_batch.clicked.connect(self.purge_from_batch)
        self.actionpreviousModel.triggered.connect(self.load_previous_rod_file_in_batch)
        self.actionnextModel.triggered.connect(self.load_next_rod_file_in_batch)

        #pushbuttons for data handeling
        self.pushButton_load_data.clicked.connect(self.load_data_ctr)
        self.pushButton_append_data.clicked.connect(self.append_data)
        self.pushButton_delete_data.clicked.connect(self.delete_data)
        self.pushButton_save_data.clicked.connect(self.save_data)
        self.pushButton_update_mask.clicked.connect(self.update_mask_info_in_data)
        self.pushButton_use_all.clicked.connect(self.use_all_data)
        self.pushButton_use_none.clicked.connect(self.use_none_data)
        self.pushButton_use_selected.clicked.connect(self.use_selected_data)
        self.pushButton_invert_use.clicked.connect(self.invert_use_data)
        self.pushButton_dummy_data.clicked.connect(self.generate_dummy_data_dialog)

        #pushbuttons for structure view
        self.pushButton_azimuth_0.clicked.connect(self.azimuth_0)
        self.pushButton_azimuth_90.clicked.connect(self.azimuth_90)
        self.pushButton_elevation_0.clicked.connect(self.elevation_0)
        self.pushButton_elevation_90.clicked.connect(self.elevation_90)
        self.pushButton_parallel.clicked.connect(self.parallel_projection)
        self.pushButton_projective.clicked.connect(self.projective_projection)
        self.pushButton_pan.clicked.connect(self.pan_msv_view)
        self.pushButton_start_spin.clicked.connect(self.start_spin)
        self.pushButton_stop_spin.clicked.connect(self.stop_spin)
        self.pushButton_xyz.clicked.connect(self.save_structure_file)

        #spinBox to save the domain_tag
        self.spinBox_domain.valueChanged.connect(self.update_domain_index)

        #pushbutton to load/save script
        self.pushButton_load_script.clicked.connect(self.load_script)
        self.pushButton_save_script.clicked.connect(self.save_script)
        # self.pushButton_modify_script.clicked.connect(self.modify_script)

        #pushbutton to load/save parameter file
        self.pushButton_load_table.clicked.connect(self.load_par)
        self.pushButton_save_table.clicked.connect(self.save_par)
        self.pushButton_remove_rows.clicked.connect(self.remove_selected_rows)
        self.pushButton_add_one_row.clicked.connect(self.append_one_row)
        self.pushButton_add_par_set.clicked.connect(lambda:self.append_par_set(par_selected=None))
        self.pushButton_add_all_pars.clicked.connect(self.append_all_par_sets)
        self.pushButton_fit_all.clicked.connect(self.fit_all)
        self.pushButton_fit_none.clicked.connect(self.fit_none)
        self.pushButton_fit_selected.clicked.connect(self.fit_selected)
        self.pushButton_fit_next_5.clicked.connect(self.fit_next_5)
        self.pushButton_invert_fit.clicked.connect(self.invert_fit)
        self.pushButton_update_pars.clicked.connect(self.update_model_parameter)
        self.horizontalSlider_par.valueChanged.connect(self.play_with_one_par)
        self.pushButton_scan.clicked.connect(self.scan_one_par)

        #pushButton to operate plots
        self.pushButton_update_plot.clicked.connect(lambda:self.update_structure_view(compile = True))
        self.pushButton_update_plot.clicked.connect(lambda:self.update_plot_data_view_upon_simulation(q_correction = False))
        self.pushButton_update_plot.clicked.connect(self.update_par_bar_during_fit)
        self.pushButton_update_plot.clicked.connect(self.update_electron_density_profile)
        self.pushButton_previous_screen.clicked.connect(self.show_plots_on_previous_screen)
        self.pushButton_next_screen.clicked.connect(self.show_plots_on_next_screen)
        #q correction widgets
        self.groupBox_q_correction.hide()
        self.pushButton_show.clicked.connect(lambda:self.update_plot_data_view_upon_simulation(q_correction = True))
        self.pushButton_append.clicked.connect(self.append_L_scale)
        self.pushButton_reset.clicked.connect(self.reset_L_scale)
        self.fit_q_correction = False
        self.apply_q_correction = False
        self.pushButton_fit.clicked.connect(self.fit_q)
        self.pushButton_apply.clicked.connect(self.update_q)
        self.pushButton_q_correction.clicked.connect(lambda:self.groupBox_q_correction.show())
        self.pushButton_hide.clicked.connect(lambda:self.groupBox_q_correction.hide())
        #select dataset in the viewer
        self.comboBox_dataset.activated.connect(self.update_data_view)

        #GAF viewer
        self.pushButton_generate_GAF.clicked.connect(self.generate_gaf_plot)

        #syntax highlight for script
        self.plainTextEdit_script.setStyleSheet("""QPlainTextEdit{
                                font-family:'Consolas';
                                font-size:14pt;
                                color: #ccc;
                                background-color: #2b2b2b;}""")
        self.plainTextEdit_script.setTabStopWidth(self.plainTextEdit_script.fontMetrics().width(' ')*4)

        #table view for parameters set to selecting row basis
        self.tableWidget_pars.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tableWidget_data.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.timer_update_structure = QtCore.QTimer(self)
        self.timer_update_structure.timeout.connect(self.pushButton_update_plot.click)
        self.timer_spin_msv = QtCore.QTimer(self)
        self.timer_spin_msv.timeout.connect(self.spin_msv)
        self.azimuth_angle = 0
        self.setup_plot()
        self._load_par()

        #widgets for plotting figures
        self.widget_fig.parent = self
        self.pushButton_extract_data.clicked.connect(lambda:self.widget_fig.extract_data_all())
        self.pushButton_reset_plot.clicked.connect(lambda:self.widget_fig.reset())
        self.pushButton_init_pars.clicked.connect(lambda:self.widget_fig.init_pandas_model())
        self.pushButton_plot_figures.clicked.connect(lambda:self.widget_fig.create_plots())
        self.pushButton_clear_plot.clicked.connect(lambda:self.widget_fig.clear_plot())

        #widgets for model result evaluation
        self.pushButton_cov.clicked.connect(self.generate_covarience_matrix)
        self.pushButton_sensitivity.clicked.connect(self.screen_parameters)

        #help tree widget
        # self.treeWidget.itemDoubleClicked.connect(self.open_help_doc)

    def generate_gaf_plot(self):
        self.simulate_model(compile = True)
        data_series = sum(self.normalized_datasets,[])
        self.widget_gaf.create_plots(serie_data = data_series, plot_type = self.comboBox_plot_channel.currentText(),relative = self.checkBox_GAF_relatively.isChecked())

    def append_L_scale(self):
        if self.lineEdit_L_container.text()=='':
            self.lineEdit_L_container.setText(self.lineEdit_L.text())
            self.lineEdit_scale_container.setText(self.lineEdit_scale.text())
        else:
            self.lineEdit_L_container.setText(self.lineEdit_L_container.text()+',{}'.format(self.lineEdit_L.text()))
            self.lineEdit_scale_container.setText(self.lineEdit_scale_container.text()+',{}'.format(self.lineEdit_scale.text()))

    def reset_L_scale(self):
        self.lineEdit_L_container.setText('')
        self.lineEdit_scale_container.setText('')

    def fit_q(self):
        L_list = eval('[{}]'.format(self.lineEdit_L_container.text()))
        scale_list = eval('[{}]'.format(self.lineEdit_scale_container.text()))
        if len(L_list)==0 or len(scale_list)==0 or len(L_list)!=len(scale_list):
            return
        lam = self.model.script_module.wal
        R_tth = float(self.lineEdit_r_tth.text())
        unitcell = self.model.script_module.unitcell
        delta,c_off,scale = fit_q_correction(lam,R_tth,L_list,scale_list, unitcell.c*np.sin(unitcell.beta))
        self.q_correction_factor = {'L_bragg':L_list[0],'delta':delta, 'c_off':c_off, 'scale':scale}
        self.fit_q_correction = True
        self.update_plot_data_view_upon_simulation(q_correction = True)
        self.fit_q_correction = False

    def update_q(self):
        reply = QtGui.QMessageBox.question(self, 'Message',
        "Are you sure to update the data with q correction results?", QtGui.QMessageBox.Yes | 
        QtGui.QMessageBox.No, QtGui.QMessageBox.No)
        if reply == QtGui.QMessageBox.Yes:
            #print("YES")
            self.apply_q_correction = True
            self.fit_q()
            self.apply_q_correction = False
        else:
            return

    def apply_q_correction_results(self, index, LL):
        self.model.data_original[index].x = LL
        self.model.data = copy.deepcopy(self.model.data_original)
        [each.apply_mask() for each in self.model.data]
        self.model.data.concatenate_all_ctr_datasets()
        self.simulate_model()

    def generate_covarience_matrix(self):
        fom_level = float(self.lineEdit_error.text())
        if len(self.run_fit.solver.optimizer.par_evals)==0:
            return
        condition = (self.run_fit.solver.optimizer.fom_evals.array()+1)<(self.run_fit.solver.model.fom+1)*(1+fom_level)
        target_matrix = self.run_fit.solver.optimizer.par_evals[condition]
        df = pd.DataFrame(target_matrix)
        corr = df.corr()
        corr.index += 1
        corr = corr.rename(columns = lambda x:str(int(x)+1))
        self.covariance_matrix = corr
        #cmap: coolwarm, plasma, hsv
        self.textEdit_cov.setHtml(corr.style.background_gradient(cmap='coolwarm').set_precision(3).render())

    #calculate the sensitivity of each fit parameter
    #sensitivity: how much percentage increase of a parameter has to be applied to achived ~10% increase in fom?
    # the increase rate of fom divided by that for par give rise to the numerical representative of sensitivity
    #In the end, all sensitivity values are normalized to have the max value equal to 1 for better comparison.
    def screen_parameters(self):
        index_fit_pars = [i for i in range(len(self.model.parameters.data)) if self.model.parameters.data[i][2]]
        #par_names = ['{}.'.format(i) for i in range(1,len(index_fit_pars)+1)]
        #print(par_names)
        epoch_list = [0]*len(index_fit_pars)
        fom_diff_list = [0]*len(index_fit_pars)
        #each epoch, increase value by 2%
        epoch_step = float(self.lineEdit_step.text())
        max_epoch = int(self.lineEdit_epoch.text())
        for i in index_fit_pars:
            par = self.model.parameters.get_value(i, 0)
            print('Screen par {} now!'.format(par))
            current_value = self.model.parameters.get_value(i, 1)
            current_fom = self.model.fom
            current_vec = copy.deepcopy(self.run_fit.solver.optimizer.best_vec)
            epoch = 0
            while epoch<max_epoch:
                epoch = epoch + 1
                #self.model.parameters.set_value(i, 1, current_value*(1+epoch_step*epoch))
                #self.model.simulate()
                current_vec[index_fit_pars.index(i)] = current_value+abs(current_value)*epoch_step*epoch
                print(epoch, current_value, abs(current_value)*epoch_step*epoch)
                fom = self.run_fit.solver.optimizer.calc_fom(current_vec)
                #offset off 1 is used just in case the best fom is very close to 0
                if (fom+1)>(current_fom+1)*(1+0.1):
                    epoch_list[index_fit_pars.index(i)] = epoch*epoch_step
                    fom_diff_list[index_fit_pars.index(i)] = (fom - current_fom)/current_fom
                    #set the original value back
                    self.model.parameters.set_value(i, 1, current_value)
                    #print(epoch_list)
                    print('Break')
                    break
                if epoch == max_epoch:
                    fom_diff_list[index_fit_pars.index(i)] = (fom - current_fom)/current_fom
                    print(fom, current_fom)
                    epoch_list[index_fit_pars.index(i)] = epoch*epoch_step
                    self.model.parameters.set_value(i, 1, current_value)

        sensitivity = np.array(fom_diff_list)/np.array(epoch_list)
        self.plot_bar_chart(sensitivity/max(sensitivity))

    def plot_bar_chart(self, data):
        self.sensitivity_data = list(data)
        par_names = [str(i+1) for i in range(len(data))]
        self.widget_sensitivity_bar.clear()
        bg1 = pg.BarGraphItem(x=list(range(1,len(data)+1)), height=data, width=0.3, brush='g')
        ax_bar = self.widget_sensitivity_bar.addPlot(clear = True)
        ax_bar.addItem(bg1)
        #[list(zip(list(range(1,len(percents)+1)),[str(each) for each in range(1,len(percents)+1)]))]
        ax_bar.getAxis('bottom').setTicks([list(zip(list(range(1,len(data)+1)),par_names))])
        ax_bar.getAxis('bottom').setLabel('parameters')
        ax_bar.getAxis('left').setLabel('Normalized sensitivity')
        ax_bar.setYRange(0, 1, padding = 0.1)
        # ax_bar.autoRange()

    def open_help_doc(self):
        print('Double clicked signal received!')
        # return self.treeWidget.currentItem()

    def generate_script_dialog(self):
        dlg = ScriptGeneraterDialog(self)
        hightlight = syntax_pars.PythonHighlighter(dlg.plainTextEdit_script.document())
        dlg.plainTextEdit_script.show()
        dlg.plainTextEdit_script.setPlainText(dlg.plainTextEdit_script.toPlainText())
        dlg.exec()

    def generate_dummy_data_dialog(self):
        dlg = DummydataGeneraterDialog(self)
        dlg.exec()

    def start_nlls(self):
        self.thread_nlls_fit = Thread(target = self.nlls_fit.fit_model, args = ())
        self.thread_nlls_fit.start()
        self.timer_nlls.start(50)

    def update_status_nlls(self):
        if not self.nlls_fit.running:
            self.timer_nlls.stop()
            self.statusbar.showMessage('Finish running model based on NLLS: fom = {} at trial_{}'.format(self.nlls_fit.fom,self.nlls_fit.run_num))
            self.update_error_bars_from_nlls()
        else:
            self.statusbar.showMessage('Running model based on NLLS: fom = {} at trial_{}'.format(self.nlls_fit.fom,self.nlls_fit.run_num))

    def update_error_bars_from_nlls(self):
        errors = self.nlls_fit.perr
        accum_fit_par = -1
        for i in range(self.tableWidget_pars.rowCount()):
            if self.tableWidget_pars.cellWidget(i,2).isChecked():
                accum_fit_par = accum_fit_par+1
                #we only change the error but the values are maintained from DE fit results
                self.tableWidget_pars.item(i,5).setText(str(round(errors[accum_fit_par],9)))

    def hook_to_batch(self):
        self.run_batch.set_hooker(True)
        self.run_batch.rod_files = []
        for i in range(self.listWidget_rod_files.count()):
            self.run_batch.rod_files.append(os.path.join(self.lineEdit_folder_of_rod_files.text(),self.listWidget_rod_files.item(i).text()))
        self.open_model_with_path(self.run_batch.rod_files[0])
        self.listWidget_rod_files.setCurrentRow(0)

    def purge_from_batch(self):
        self.run_batch.set_hooker(False)
        self.run_batch.rod_files = []

    def load_rod_files(self):
        '''
        load all rod files(*.rod) located in a selected folder
        '''
        folder = str(QFileDialog.getExistingDirectory(self, "Select Directory"))
        self.lineEdit_folder_of_rod_files.setText(folder)
        for file in os.listdir(folder):
            if file.endswith('.rod'):
                self.listWidget_rod_files.addItem(file)

    def remove_rod_files(self):
        '''
        remove all files in the list
        '''
        self.listWidget_rod_files.clear()

    def remove_selected_rod_files(self):
        items = self.listWidget_rod_files.selectedItems()
        if not items:return
        for item in items:
            self.listWidget_rod_files.takeItem(self.listWidget_rod_files.row(item))

    def load_next_rod_file_in_batch(self):
        if not hasattr(self.run_batch,'rod_files'):
            pass
        else:
            #if you are now at the last rod file, then roll back to the first one
            if self.run_batch.rod_files.index(self.rod_file)==(len(self.run_batch.rod_files)-1):
                self.open_model_with_path(self.run_batch.rod_files[0])
                self.listWidget_rod_files.setCurrentRow(0)
            else:
                target_index = self.run_batch.rod_files.index(self.rod_file) + 1
                self.open_model_with_path(self.run_batch.rod_files[target_index])
                self.listWidget_rod_files.setCurrentRow(target_index)

    def load_previous_rod_file_in_batch(self):
        if not hasattr(self.run_batch,'rod_files'):
            pass
        else:
            #if you are now at the first rod file, then roll forward to the last one
            if self.run_batch.rod_files.index(self.rod_file)==0:
                self.open_model_with_path(self.run_batch.rod_files[-1])
                self.listWidget_rod_files.setCurrentRow(len(self.run_batch.rod_files)-1)
            else:
                target_index = self.run_batch.rod_files.index(self.rod_file) - 1
                self.open_model_with_path(self.run_batch.rod_files[target_index])
                self.listWidget_rod_files.setCurrentRow(target_index)

    def show_plots_on_next_screen(self):
        """
        show plots on next screen, if one screen is not enough to fill all plots
        """
        if not hasattr(self,"num_screens_plot"):
            return

        if self.num_screens_plot>1:
            if self.current_index_plot_screen<(self.num_screens_plot-1):
                self.update_plot_dimension(self.current_index_plot_screen+1)
                self.update_plot_data_view()
            else:
                pass
        else:
            pass

    def show_plots_on_previous_screen(self):
        """
        show plots on previous screen
        """

        if not hasattr(self,"num_screens_plot"):
            return

        if self.num_screens_plot>1:
            if self.current_index_plot_screen>0:
                self.update_plot_dimension(self.current_index_plot_screen-1)
                self.update_plot_data_view()
            else:
                pass
        else:
            pass

    def toggle_data_panel(self):
        """data panel on the left side of GUI main frame"""
        self.tabWidget_data.setVisible(self.actionData.isChecked())

    def toggle_plot_panel(self):
        """plot panel on the top right side of main GUI frame"""
        self.tabWidget.setVisible(self.actionPlot.isChecked())

    def toggle_script_panel(self):
        """script panel on the bottom right side of main GUI frame"""
        self.tabWidget_2.setVisible(self.actionScript.isChecked())

    def update_domain_index(self):
        """update domain index, triggering the associated structure to show"""
        self.domain_tag = int(self.spinBox_domain.text())
        if self.model.compiled:
            self.widget_edp.items = []
            # self.widget_msv_top.items = []
            self.init_structure_view()
        else:
            pass

    def parallel_projection(self):
        self.widget_edp.opts['distance'] = 2000
        self.widget_edp.opts['fov'] = 1
        # self.widget_msv_top.opts['distance'] = 2000
        # self.widget_msv_top.opts['fov'] = 1
        self.update_structure_view()

    def projective_projection(self):
        self.widget_edp.opts['distance'] = 25
        self.widget_edp.opts['fov'] = 60
        # self.widget_msv_top.opts['distance'] = 25
        # self.widget_msv_top.opts['fov'] = 60
        self.update_structure_view()

    def pan_msv_view(self):
        value = int(self.spinBox_pan_pixel.text())
        self.widget_edp.pan(value*int(self.checkBox_x.isChecked()),value*int(self.checkBox_y.isChecked()),value*int(self.checkBox_z.isChecked()))

    def update_camera_position(self,widget_name = 'widget_edp', angle_type="azimuth", angle=0):
        getattr(self,widget_name).setCameraPosition(pos=None, distance=None, \
            elevation=[None,angle][int(angle_type=="elevation")], \
                azimuth=[None,angle][int(angle_type=="azimuth")])

    def azimuth_0(self):
        self.update_camera_position(angle_type="azimuth", angle=0)

    def azimuth_90(self):
        self.update_camera_position(angle_type="azimuth", angle=90)

    def start_spin(self):
        self.timer_spin_msv.start(100)

    def stop_spin(self):
        self.timer_spin_msv.stop()

    def spin_msv(self):
        #if self.azimuth > 360:
            
        self.update_camera_position(angle_type="azimuth", angle=self.azimuth_angle)
        self.azimuth_angle = self.azimuth_angle + 1


    def elevation_0(self):
        self.update_camera_position(angle_type="elevation", angle=0)

    def elevation_90(self):
        self.update_camera_position(angle_type="elevation", angle=90)

    #do this after model is loaded, so that you know len(data)
    def update_plot_dimension(self, current_index_plot_screen = 0):
        """Setting the layout of data profiles"""
        def _get_index(index_in_use):
            use_or_not = []
            for i in range(len(self.model.data)):
                if self.tableWidget_data.cellWidget(i,2).isChecked():
                    use_or_not.append(True)
                else:
                    use_or_not.append(False)
            index_in_sequence = index_in_use
            total = -1
            for i, each in enumerate(use_or_not):
                if total<index_in_use:
                    if not each:
                        index_in_sequence += 1
                    else:
                        total += 1
            return index_in_sequence
        sizeObject = QtWidgets.QDesktopWidget().screenGeometry(-1)
        height, width = sizeObject.height()*25.4/dpi,sizeObject.width()*25.4/dpi
        #maximum number of plots allowd to be fit in one screen
        #assuming the minimum plot panel has a size of (w:50mm * h:40mm)
        plot_cols, plot_rows = int(width/50), int(height/40)
        self.max_num_plots_per_screen = plot_cols*plot_rows

        self.widget_data.clear()
        self.widget_data.ci.currentRow = 0
        self.widget_data.ci.currentCol = 0

        self.data_profiles = []
        #only consider the data in use
        total_datasets = len([1 for i in range(len(self.model.data)) if self.tableWidget_data.cellWidget(i,2).isChecked()])
        #total_datasets = len([each for each in self.model.data if each.use])

        if total_datasets<self.max_num_plots_per_screen:
            self.num_screens_plot = 1
        else:
            self.num_screens_plot = int(total_datasets/self.max_num_plots_per_screen)+[0,1][int((total_datasets%self.max_num_plots_per_screen)>0)]
        self.current_index_plot_screen = current_index_plot_screen
        if self.num_screens_plot>1:#more than one screen
            if self.current_index_plot_screen<(self.num_screens_plot-1):
                columns = plot_cols#should be occupied in maximum
                num_plots_on_current_screen = self.max_num_plots_per_screen
            else:#last screen
                num_plots_ = total_datasets%self.max_num_plots_per_screen
                if num_plots_ == 0:
                    columns = plot_cols
                    num_plots_on_current_screen = self.max_num_plots_per_screen
                else:
                    num_plots_on_current_screen = num_plots_
                    if num_plots_>10:
                        columns = 4
                    else:
                        columns = 2
        elif self.num_screens_plot==1:#only one screen
            if total_datasets==self.max_num_plots_per_screen:
                num_plots_on_current_screen = self.max_num_plots_per_screen
                columns = plot_cols
            else:
                num_plots_on_current_screen = total_datasets
                if total_datasets>10:
                    columns = 4
                else:
                    columns = 2

        #current list of ax handle
        self.num_plots_on_current_screen = num_plots_on_current_screen
        offset = self.current_index_plot_screen*self.max_num_plots_per_screen
        for i in range(num_plots_on_current_screen):
            if 1:
                hk_label = '{}{}_{}'.format(str(int(self.model.data[_get_index(i+offset)].extra_data['h'][0])),str(int(self.model.data[_get_index(i+offset)].extra_data['k'][0])),str(self.model.data[_get_index(i+offset)].extra_data['Y'][0]))
                if (i%columns)==0 and (i!=0):
                    self.widget_data.nextRow()
                    self.data_profiles.append(self.widget_data.addPlot(title=hk_label))
                else:
                    self.data_profiles.append(self.widget_data.addPlot(title=hk_label))

    def setup_plot(self):
        self.fom_evolution_profile = self.widget_fom.addPlot()
        self.par_profile = self.widget_pars.addPlot()
        self.fom_scan_profile = self.widget_fom_scan.addPlot()
        self.fom_scan_profile.getAxis('left').setLabel('Electron denstiy (per water)')
        self.fom_scan_profile.getAxis('bottom').setLabel('Height (Å)')

    def update_data_check_attr(self):
        """update the checkable attr of each dataset: use, show, showerror"""
        re_simulate = False
        for i in range(len(self.model.data)):
            #model.data: masked data
            self.model.data[i].show = self.tableWidget_data.cellWidget(i,1).isChecked()
            self.model.data[i].use_error = self.tableWidget_data.cellWidget(i,3).isChecked()
            #model.data_original: unmasked data for model saving later
            self.model.data_original[i].show = self.tableWidget_data.cellWidget(i,1).isChecked()
            self.model.data_original[i].use_error = self.tableWidget_data.cellWidget(i,3).isChecked()
            if self.model.data[i].use!=self.tableWidget_data.cellWidget(i,2).isChecked():
                re_simulate = True
                self.model.data[i].use = self.tableWidget_data.cellWidget(i,2).isChecked()
                self.model.data_original[i].use = self.tableWidget_data.cellWidget(i,2).isChecked()
        if re_simulate:
            self.simulate_model()

    def calc_f_ideal(self):
        self.f_ideal = []
        for i in range(len(self.model.data)):
            each = self.model.data[i]
            if each.x[0]>1000:#indicate energy column
                self.f_ideal.append(self.model.script_module.sample.calc_f_ideal(each.extra_data['h'], each.extra_data['k'], each.extra_data['Y'])**2)
            else:
                self.f_ideal.append(self.model.script_module.sample.calc_f_ideal(each.extra_data['h'], each.extra_data['k'], each.x)**2)

    def update_plot_data_view(self):
        """update views of all figures if script is compiled, while only plot data profiles if otherwise"""
        def _get_index(index_in_use):
            index_in_sequence = index_in_use
            total = -1
            for i, each in enumerate(self.model.data):
                if total<index_in_use:
                    if not each.use:
                        index_in_sequence += 1
                    else:
                        total += 1
            return index_in_sequence

        if self.model.compiled:
            self.update_data_check_attr()
            self.update_plot_data_view_upon_simulation()
            self.update_electron_density_profile()
        else:
            offset = self.max_num_plots_per_screen*self.current_index_plot_screen
            for i in range(self.num_plots_on_current_screen):
                fmt = self.tableWidget_data.item(i+offset,4).text()
                fmt_symbol = list(fmt.rstrip().rsplit(';')[0].rsplit(':')[1])
                self.data_profiles[i].plot(self.model.data[_get_index(i+offset)].x, self.model.data[_get_index(i+offset)].y,pen = None,  symbolBrush=fmt_symbol[1], symbolSize=int(fmt_symbol[0]),symbolPen=fmt_symbol[2], clear = True)
            [each.setLogMode(x=False,y=self.tableWidget_data.cellWidget(_get_index(self.data_profiles.index(each)),1).isChecked()) for each in self.data_profiles]
            [each.autoRange() for each in self.data_profiles]

    def update_electron_density_profile(self):
        if self.lineEdit_z_min.text()!='':
            z_min = float(self.lineEdit_z_min.text())
        else:
            z_min = -20
        if self.lineEdit_z_max.text()!='':
            z_max = float(self.lineEdit_z_max.text())
        else:
            z_max = 100
        raxs_A_list, raxs_P_list = [], []
        #num_raxs = len(self.model.data)-1
        #items for raxs dates have value >=100 in the data_sequence attribute
        num_raxs = sum(np.array(self.model.data.data_sequence)>=100)
        if hasattr(self.model.script_module, "rgh_raxs"):
            for i in range(num_raxs):
                raxs_A_list.append(eval("self.model.script_module.rgh_raxs.getA_{}()".format(i+1)))
                raxs_P_list.append(eval("self.model.script_module.rgh_raxs.getP_{}()".format(i+1)))
        else:
            raxs_A_list.append(0)
            raxs_P_list.append(0)
        # raxs_A_list = raxs_A_list[0:2]
        # raxs_P_list = raxs_P_list[0:2]
        HKL_raxs_list = [[],[],[]]
        for each in self.model.data:
            if each.x[0]>=100:
                HKL_raxs_list[0].append(each.extra_data['h'][0])
                HKL_raxs_list[1].append(each.extra_data['k'][0])
                HKL_raxs_list[2].append(each.extra_data['Y'][0])
        # HKL_raxs_list = [HKL_raxs_list[0][0:2],HKL_raxs_list[1][0:2],HKL_raxs_list[2][0:2]]
        if hasattr(self.model.script_module, "RAXS_EL"):
            raxs_el = getattr(self.model.script_module, "RAXS_EL")
        else:
            raxs_el = None
        try:
            if self.run_fit.running or self.run_batch.running:
                #if model is running, disable showing e profile
                pass
            else:
                self.fom_scan_profile.addLegend(offset = (-10,20))
                label,edf = self.model.script_module.sample.plot_electron_density_superrod(z_min=z_min, z_max=z_max,N_layered_water=500,resolution =1000, raxs_el = raxs_el, use_sym = self.checkBox_symmetry.isChecked())
                #here only plot the total electron density of domain specified by domain_tag
                domain_tag = int(self.spinBox_domain.text())
                self.fom_scan_profile.plot(edf[domain_tag][0],edf[domain_tag][1],pen = {'color': "w", 'width': 1},clear = True)
                self.fom_scan_profile.plot(edf[domain_tag][0],edf[domain_tag][1],fillLevel=0, brush = (0,200,0,100),clear = False, name = 'Total ED')
                if len(edf[domain_tag])==4:
                    self.fom_scan_profile.plot(edf[domain_tag][0],edf[domain_tag][2],fillLevel=0, brush = (200,0,0,80),clear = False)
                    self.fom_scan_profile.plot(edf[domain_tag][0],edf[domain_tag][3],fillLevel=0, brush = (0,0,250,80),clear = False, name = 'Water Layer')
                if hasattr(self.model.script_module, "rgh_raxs"):
                    # print(HKL_raxs_list)
                    # print(raxs_P_list) 
                    # print(raxs_A_list)
                    # z_plot,eden_plot,_=self.model.script_module.sample.fourier_synthesis(np.array(HKL_raxs_list),np.array(raxs_P_list).transpose(),np.array(raxs_A_list).transpose(),z_min=z_min,z_max=z_max,resonant_el=self.model.script_module.raxr_el,resolution=1000,water_scaling=0.33)
                    z_plot,eden_plot,_=self.model.script_module.sample.fourier_synthesis(np.array(HKL_raxs_list),np.array(raxs_P_list).transpose(),np.array(raxs_A_list).transpose(),z_min=z_min,z_max=z_max,resonant_el=self.model.script_module.RAXS_EL,resolution=1000,water_scaling=0.33)
                    self.fom_scan_profile.plot(z_plot,eden_plot,fillLevel=0, brush = (200,0,200,100),clear = False, name = 'ED based on Fourier Synthesis')
                self.fom_scan_profile.autoRange()
        except:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to draw e density profile!')
            logging.getLogger().exception('Fatal error encountered during drawing e density profile!')
            self.tabWidget_data.setCurrentIndex(6)

    def update_plot_data_view_upon_simulation(self, q_correction = False):
        self.normalized_datasets = []
        def _get_index(index_in_use):
            index_in_sequence = index_in_use
            total = -1
            for i, each in enumerate(self.model.data):
                if total<index_in_use:
                    if not each.use:
                        index_in_sequence += 1
                    else:
                        total += 1
            return index_in_sequence

        offset = self.max_num_plots_per_screen*self.current_index_plot_screen
        for i in range(self.num_plots_on_current_screen):
            if 1:
                #plot ideal structure factor
                f_ideal = 1
                scale_factor = 1
                rod_factor = 1
                try:
                    specular_condition = int(round(self.model.data[_get_index(i+offset)].extra_data['h'][0],0))==0 and int(round(self.model.data[_get_index(i+offset)].extra_data['k'][0],0))==0
                    scale_factor = [self.model.script_module.rgh.scale_nonspecular_rods,self.model.script_module.rgh.scale_specular_rod][int(specular_condition)]
                    h_, k_ = int(round(self.model.data[_get_index(i+offset)].extra_data['h'][0],0)),int(round(self.model.data[_get_index(i+offset)].extra_data['k'][0],0))
                    extra_scale_factor = 'scale_factor_{}{}L'.format(h_,k_)
                    if hasattr(self.model.script_module.rgh,extra_scale_factor):
                        rod_factor = getattr(self.model.script_module.rgh, extra_scale_factor)
                    else:
                        rod_factor = 1
                    if self.checkBox_norm.isChecked():
                        f_ideal = self.f_ideal[_get_index(i+offset)]*scale_factor*rod_factor
                    else:
                        self.data_profiles[i].plot(self.model.data[_get_index(i+offset)].x, self.f_ideal[_get_index(i+offset)]*scale_factor*rod_factor,pen = {'color': "w", 'width': 1},clear = True)
                except:
                    pass

                fmt = self.tableWidget_data.item(i+offset,4).text()
                fmt_symbol = list(fmt.rstrip().rsplit(';')[0].rsplit(':')[1])
                line_symbol = list(fmt.rstrip().rsplit(';')[1].rsplit(':')[1])
                if not q_correction:
                    self.data_profiles[i].plot(self.model.data[_get_index(i+offset)].x, self.model.data[_get_index(i+offset)].y/f_ideal,pen = None,  symbolBrush=fmt_symbol[1], symbolSize=int(fmt_symbol[0]),symbolPen=fmt_symbol[2],clear = False)
                else:
                    if self.model.data[_get_index(i+offset)].name == self.comboBox_dataset2.currentText():
                        L_q_correction = self.model.data_original[_get_index(i+offset)].x
                        data_q_correction = self.model.data_original[_get_index(i+offset)].y
                        unitcell = self.model.script_module.unitcell
                        cell = [unitcell.a, unitcell.b, unitcell.c, unitcell.alpha, unitcell.beta, unitcell.gamma]
                        lam = self.model.script_module.wal
                        scale = float(self.lineEdit_scale.text())
                        current_L = int(self.lineEdit_L.text())
                        if not self.fit_q_correction:
                            LL, new_data= q_correction_for_one_Bragg_peak(L = L_q_correction,data = data_q_correction, cell = cell, lam = lam, L_bragg = current_L, scale=scale,delta=0,c_off=0, R_tth = float(self.lineEdit_r_tth.text()))
                        else:
                            LL, new_data= q_correction_for_one_rod(L = L_q_correction, data = data_q_correction, cell = cell, lam = lam, correction_factor_dict = self.q_correction_factor, R_tth = float(self.lineEdit_r_tth.text()))
                            if self.apply_q_correction:
                                self.apply_q_correction_results(_get_index(i+offset), LL)
                        # print(data)
                        # print(scale_factor*rod_factor)
                        #recalculate f_ideal
                        #self.model.script_module.unitcell.set_c(new_c)
                        #self.calc_f_ideal()
                        #f_ideal = self.f_ideal[_get_index(i+offset)]*scale_factor*rod_factor
                        self.data_profiles[i].plot(LL, new_data, pen = None,  symbolBrush=fmt_symbol[1], symbolSize=int(fmt_symbol[0]),symbolPen=fmt_symbol[2],clear = True)
                if self.tableWidget_data.cellWidget(_get_index(i+offset),3).isChecked():
                    #create error bar data, graphiclayout widget doesn't have a handy api to plot lines along with error bars in a log scale
                    #disable this while the model is running
                    if not self.run_fit.solver.optimizer.running:
                        '''#this solution does not work in a log scale
                        x, y, error = self.model.data[_get_index(i+offset)].x, self.model.data[_get_index(i+offset)].y, self.model.data[_get_index(i+offset)].error/2
                        err = pg.ErrorBarItem(x=x, y=y, top=error, bottom=error)
                        self.data_profiles[i].addItem(err)
                        '''
                        x = np.append(self.model.data[_get_index(i+offset)].x[:,np.newaxis],self.model.data[_get_index(i+offset)].x[:,np.newaxis],axis=1)
                        y_d = self.model.data[_get_index(i+offset)].y[:,np.newaxis] - self.model.data[_get_index(i+offset)].error[:,np.newaxis]/2
                        y_u = self.model.data[_get_index(i+offset)].y[:,np.newaxis] + self.model.data[_get_index(i+offset)].error[:,np.newaxis]/2
                        y = np.append(y_d,y_u,axis = 1)
                        for ii in range(len(y)):
                            self.data_profiles[i].plot(x=x[ii],y=y[ii],pen={'color':'w', 'width':1},clear = False)
                

                #plot simulated results
                if not q_correction:
                    if self.tableWidget_data.cellWidget(_get_index(i+offset),2).isChecked():
                        self.data_profiles[i].plot(self.model.data[_get_index(i+offset)].x, self.model.data[_get_index(i+offset)].y_sim/f_ideal,pen={'color': line_symbol[1], 'width': int(line_symbol[0])},  clear = False)
                        # self.normalized_datasets.append(list(np.log10(self.model.data[_get_index(i+offset)].y_sim/f_ideal)))
                        self.normalized_datasets.append(list(self.model.data[_get_index(i+offset)].y_sim/f_ideal))
                    else:
                        pass
        [each.setLogMode(x=False,y=self.tableWidget_data.cellWidget(_get_index(self.data_profiles.index(each)+offset),1).isChecked()) for each in self.data_profiles]
        [each.autoRange() for each in self.data_profiles]
        fom_log = np.array(self.run_fit.solver.optimizer.fom_log)
        self.fom_evolution_profile.plot(fom_log[:,0],fom_log[:,1],pen={'color': 'r', 'width': 2}, clear = True)
        self.fom_evolution_profile.autoRange()
        
    def update_par_bar_during_fit(self):
        """update bar chart during fit, which tells the current best fit and the searching range of each fit parameter"""
        if self.run_fit.running or self.run_batch.running:
            if self.run_fit.running:
                par_max = self.run_fit.solver.optimizer.par_max
                par_min = self.run_fit.solver.optimizer.par_min
                vec_best = copy.deepcopy(self.run_fit.solver.optimizer.best_vec)
                vec_best = (vec_best-par_min)/(par_max-par_min)
                pop_vec = np.array(copy.deepcopy(self.run_fit.solver.optimizer.pop_vec))
            elif self.run_batch.running:
                par_max = self.run_batch.solver.optimizer.par_max
                par_min = self.run_batch.solver.optimizer.par_min
                vec_best = copy.deepcopy(self.run_batch.solver.optimizer.best_vec)
                vec_best = (vec_best-par_min)/(par_max-par_min)
                pop_vec = np.array(copy.deepcopy(self.run_batch.solver.optimizer.pop_vec))

            trial_vec_min =[]
            trial_vec_max =[]
            for i in range(len(par_max)):
                trial_vec_min.append((np.min(pop_vec[:,i])-par_min[i])/(par_max[i]-par_min[i]))
                trial_vec_max.append((np.max(pop_vec[:,i])-par_min[i])/(par_max[i]-par_min[i]))
            trial_vec_min = np.array(trial_vec_min)
            trial_vec_max = np.array(trial_vec_max)
            bg = pg.BarGraphItem(x=range(len(vec_best)), y=(trial_vec_max + trial_vec_min)/2, height=(trial_vec_max - trial_vec_min)/2, brush='b', width = 0.8)
            self.par_profile.clear()
            self.par_profile.addItem(bg)
            self.par_profile.plot(vec_best, pen=(0,0,0), symbolBrush=(255,0,0), symbolPen='w')
        else:
            pass

    def calculate_error_bars(self):
        """
        cal the error bar for each fit par after fit is completed
        note the error bar values are only estimated from all intermediate fit reuslts from all fit generations,
        and the error may not accutely represent the statistic errors. If you want to get statistical errors of 
        each fit parameter, you can run a further NLLS fit using the the best fit parameters, which is not implemented in the program.
        """
        try:
            try:
                error_bars = self.run_fit.solver.CalcErrorBars()
            except:
                try:
                    error_bars = self.run_batch.solver.CalcErrorBars()
                except:
                    return
            total_num_par = len(self.model.parameters.data)
            index_list = [i for i in range(total_num_par) if self.model.parameters.data[i][2]]
            for i in range(len(error_bars)):
                self.model.parameters.data[index_list[i]][-2] = error_bars[i]
            self.update_par_upon_load()
        except diffev.ErrorBarError as e:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to calculate error bar!')
            logging.getLogger().exception('Fatal error encountered during error calculation!')
            self.tabWidget_data.setCurrentIndex(4)
            _ = QMessageBox.question(self, 'Runtime error message', str(e), QMessageBox.Ok)


    def init_new_model(self):
        reply = QMessageBox.question(self, 'Message', 'Would you like to save the current model first?', QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if reply == QMessageBox.Yes:
            self.save_model()
        try:
            self.model = model.Model()
            self.run_fit.solver.model = self.model
            self.tableWidget_data.setRowCount(0)
            self.tableWidget_pars.setRowCount(0)
            self.plainTextEdit_script.setPlainText('')
            self.comboBox_dataset.clear()
            self.tableWidget_data_view.setRowCount(0)
            # self.update_plot_data_view()
            self._load_par()
        except Exception:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to init a new model!')
            logging.getLogger().exception('Fatal error encountered during model initiation!')
            self.tabWidget_data.setCurrentIndex(4)

    def open_model_with_path(self,path):
        fileName = path
        load_add_ = 'success'
        self.rod_file = fileName
        if fileName:
            try:
                self.setWindowTitle('Data analysis factory: CTR data modeling-->{}'.format(fileName))
                self.model.load(fileName)
                # self.load_addition()
                try:
                    self.load_addition()
                except:
                    load_add_ = 'failure'
                #add a mask attribute to each dataset
                for each in self.model.data_original:
                    if not hasattr(each,'mask'):
                        each.mask = np.array([True]*len(each.x))
                for each in self.model.data:
                    if not hasattr(each,'mask'):
                        each.mask = np.array([True]*len(each.x))
                #add model space to terminal
                self.widget_terminal.update_name_space("model",self.model)
                self.widget_terminal.update_name_space("solver",self.run_fit.solver)
                self.widget_terminal.update_name_space("win",self)

                #remove items in the msv and re-initialize it
                self.widget_edp.items = []
                # self.widget_msv_top.items = []
                #update other pars
                self.update_table_widget_data()
                self.update_plot_dimension()
                self.update_combo_box_dataset()
                self.update_plot_data_view()
                self.update_par_upon_load()
                self.update_script_upon_load()
                #model is simulated at the end of next step
                self.init_mask_info_in_data_upon_loading_model()
                #add name space for cal bond distance after simulation
                try:
                    self.widget_terminal.update_name_space("report_distance",self.model.script_module.sample.inter_atom_distance_report)
                except:
                    pass
                #now set the comboBox for par set
                self.update_combo_box_list_par_set()

                self.statusbar.clearMessage()
                self.statusbar.showMessage("Model is loaded, and {} in config loading".format(load_add_))
                # self.update_mask_info_in_data()
            except Exception:

                self.statusbar.clearMessage()
                self.statusbar.showMessage('Failure to open a model file!')
                logging.getLogger().exception('Fatal error encountered during openning a model file!')
                self.tabWidget_data.setCurrentIndex(4)

    def open_model(self):
        """open a saved model file(*.rod), which is a compressed file containing data, script and fit parameters in one place"""
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","rod file (*.rod);;zip Files (*.rar)", options=options)
        self.open_model_with_path(fileName)
        self.lineEdit_folder_of_rod_files.setText(os.path.dirname(fileName))
        self.listWidget_rod_files.addItem(os.path.basename(fileName))

    def open_model_selected_in_listWidget(self):
        """
        open a saved model file(*.rod), which is a compressed file containing data, script and fit parameters in one place
        The rod file the the selected rod file in the listWidget
        """
        fileName = os.path.join(self.lineEdit_folder_of_rod_files.text(),self.listWidget_rod_files.currentItem().text())
        self.open_model_with_path(fileName)

    def update_combo_box_list_par_set(self):
        """atomgroup and uservars instances defined in script will be colleced and displayed in this combo box"""
        attrs = self.model.script_module.__dir__()
        attr_wanted = [each for each in attrs if type(getattr(self.model.script_module, each)) in [AtomGroup, UserVars]]
        num_items = self.comboBox_register_par_set.count()
        for i in range(num_items):
            self.comboBox_register_par_set.removeItem(0)
        self.comboBox_register_par_set.insertItems(0,attr_wanted)

    def append_all_par_sets(self):
        """append fit parameters for all parset listed in the combo box, handy tool to save manual adding them in par table"""
        if "table_container" in self.model.script_module.__dir__():
            if len(self.model.script_module.table_container)!=0:
                table = self.model.script_module.table_container[::-1]
                rows = self.tableWidget_pars.selectionModel().selectedRows()
                if len(rows) == 0:
                    row_index = self.tableWidget_pars.rowCount()
                else:
                    row_index = rows[-1].row()
                for ii in range(len(table)):
                    self.tableWidget_pars.insertRow(row_index)
                    for i in range(6):
                        if i==2:
                            check_box = QCheckBox()
                            check_box.setChecked(eval(table[ii][i]))
                            self.tableWidget_pars.setCellWidget(row_index,2,check_box)
                        else:
                            if i == 0:
                                qtablewidget = QTableWidgetItem(table[ii][i])
                                qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                            elif i in [1]:
                                qtablewidget = QTableWidgetItem(table[ii][i])
                                qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                            elif i ==5:
                                qtablewidget = QTableWidgetItem('(0,0)')
                            else:
                                qtablewidget = QTableWidgetItem(table[ii][i])
                            self.tableWidget_pars.setItem(row_index,i,qtablewidget)
                self.append_one_row_at_the_end()
        else:
            par_all = [self.comboBox_register_par_set.itemText(i) for i in range(self.comboBox_register_par_set.count())]
            for par in par_all:
                self.append_par_set(par)
        self.tableWidget_pars.resizeColumnsToContents()
        self.tableWidget_pars.resizeRowsToContents()

    def append_par_set(self, par_selected = None):
        #boundary mapping for quick setting the bounds of fit pars
        bounds_map = {"setR":[0.8,1.8],"setScale":[0,1],("setdx","sorbate"):[-0.5,0.5],\
                     ("setdy","sorbate"):[-0.5,0.5],("setdz","sorbate"):[-0.1,1],("setoc","sorbate"):[0.5,3],\
                     ("setdx","surface"):[-0.1,0.1],("setdy","surface"):[-0.1,0.1],("setdz","surface"):[-0.1,0.1],\
                     ("setoc","surface"):[0.6,1],"setDelta":[-20,60],"setGamma":[0,180],"setBeta":[0,0.1]}
        def _get_bounds(attr_head,attr_item):
            for key in bounds_map:
                if type(key)==str:
                    if key in attr_item:
                        return bounds_map[key]
                else:
                    if (key[0] in attr_item) and (key[1] in attr_head):
                        return bounds_map[key]
            return []
        if par_selected==None:
            par_selected = self.comboBox_register_par_set.currentText()
        else:
            pass
        attrs = eval("self.model.script_module.{}.__dir__()".format(par_selected))
        attrs_wanted = [each for each in attrs if each.startswith("set")][::-1]

        rows = self.tableWidget_pars.selectionModel().selectedRows()
        if len(rows) == 0:
            row_index = self.tableWidget_pars.rowCount()
        else:
            row_index = rows[-1].row()
        for ii in range(len(attrs_wanted)):
            self.tableWidget_pars.insertRow(row_index)
            # current_value = eval("self.model.script_module."+par_selected+'.'+attrs_wanted[ii].replace('set','get')+"()")
            attr_temp = list(attrs_wanted[ii])
            attr_temp[0] = 'g'#set replaced by get this way
            current_value = eval("self.model.script_module."+par_selected+'.'+''.join(attr_temp)+"()")
            bounds_temp = _get_bounds(par_selected,attrs_wanted[ii])
            #update the bounds if the current value is out of the bound range
            if len(bounds_temp)==2:
                if current_value<bounds_temp[0]:
                    bounds_temp[0] = current_value
                if current_value>bounds_temp[1]:
                    bounds_temp[1] = current_value
            for i in range(6):
                if i==2:
                    check_box = QCheckBox()
                    check_box.setChecked(True)
                    self.tableWidget_pars.setCellWidget(row_index,2,check_box)
                else:
                    if i == 0:
                        qtablewidget = QTableWidgetItem(".".join([par_selected,attrs_wanted[ii]]))
                        qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                    elif i in [1]:
                        qtablewidget = QTableWidgetItem(str(round(current_value,4)))
                        qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                    elif i ==5:
                        qtablewidget = QTableWidgetItem('(0,0)')
                    elif i ==3:
                        #left boundary of fit parameter
                        if len(bounds_temp)!=0:
                            qtablewidget = QTableWidgetItem(str(round(bounds_temp[0],4)))
                        else:
                            qtablewidget = QTableWidgetItem(str(round(current_value*0.5,4)))
                    elif i ==4:
                        #right boundary of fit parameter
                        if len(bounds_temp)!=0:
                            qtablewidget = QTableWidgetItem(str(round(bounds_temp[1],4)))
                        else:
                            qtablewidget = QTableWidgetItem(str(round(current_value*1.5,4)))

                    self.tableWidget_pars.setItem(row_index,i,qtablewidget)
        self.append_one_row_at_the_end()
        self.tableWidget_pars.resizeColumnsToContents()
        self.tableWidget_pars.resizeRowsToContents()

    def auto_save_model(self):
        """model will be saved automatically during fit, for which you need to set the interval generations for saving automatically"""
        #the model will be renamed this way
        self.update_par_upon_change()
        path = self.rod_file.replace(".rod","_ran.rod")
        if path:
            #update the error bar
            self.calculate_error_bars()
            self.model.script = (self.plainTextEdit_script.toPlainText())
            self.model.save(path)
            save_add_ = 'success'
            try:
                self.save_addition()
            except:
                save_add_ = "failure"
            self.statusbar.clearMessage()
            self.statusbar.showMessage("Model is saved, and {} in config saving".format(save_add_))

    def save_model(self):
        """model will be saved automatically during fit, for which you need to set the interval generations for saving automatically"""
        #the model will be renamed this way
        try:
            path = self.rod_file
            try:
                self.calculate_error_bars()
            except:
                pass
            #self.model.script = (self.plainTextEdit_script.toPlainText())
            self.update_data_check_attr()
            self.update_par_upon_change()
            self.model.script = (self.plainTextEdit_script.toPlainText())
            self.widget_solver.update_parameter_in_solver(self)
            self.model.save(path)
            save_add_ = 'success'
            try:
                self.save_addition()
            except:
                save_add_ = "failure"
            self.statusbar.clearMessage()
            self.statusbar.showMessage("Model is saved, and {} in config saving".format(save_add_))
        except Exception:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to save model!')
            logging.getLogger().exception('Fatal error encountered during model save!')
            self.tabWidget_data.setCurrentIndex(4)

    def save_model_as(self):
        """save model file, promting a dialog widget to ask the file name to save model"""
        path, _ = QFileDialog.getSaveFileName(self, "Save file as", "", "rod file (*.rod);;zip files (*.rar)")
        if path:
            #update the rod_file attribute
            self.rod_file = path
            try:
                self.calculate_error_bars()
            except:
                pass
            #self.model.script = (self.plainTextEdit_script.toPlainText())
            self.update_data_check_attr()
            self.update_par_upon_change()
            self.model.script = (self.plainTextEdit_script.toPlainText())
            self.widget_solver.update_parameter_in_solver(self)
            self.model.save(path)
            save_add_ = 'success'
            try:
                self.save_addition()
            except:
                save_add_ = "failure"
            self.statusbar.clearMessage()
            self.statusbar.showMessage("Model is saved, and {} in config saving".format(save_add_))
            self.setWindowTitle('Data analysis factory: CTR data modeling-->{}'.format(path))

    #here save also the config pars for diffev solver
    def save_addition(self):
        """save solver parameters, pulling from pyqtgraphy.parameter_tree widget"""
        values=\
                [self.widget_solver.par.param('Diff.Ev.').param('k_m').value(),
                self.widget_solver.par.param('Diff.Ev.').param('k_r').value(),
                self.widget_solver.par.param('Diff.Ev.').param('Method').value(),
                self.widget_solver.par.param('FOM').param('Figure of merit').value(),
                self.widget_solver.par.param('FOM').param('Auto save, interval').value(),
                self.widget_solver.par.param('FOM').param('weighting factor').value(),
                self.widget_solver.par.param('FOM').param('weighting region').value(),
                self.widget_solver.par.param('Fitting').param('start guess').value(),
                self.widget_solver.par.param('Fitting').param('Generation size').value(),
                self.widget_solver.par.param('Fitting').param('Population size').value()]
        pars = ['k_m','k_r','Method','Figure of merit','Auto save, interval','weighting factor','weighting region','start guess','Generation size','Population size']
        for i in range(len(pars)):
            self.model.save_addition(pars[i],str(values[i]))
        model_info = ''
        if hasattr(self,'textEdit_note'):
            model_info = self.textEdit_note.toPlainText()
        self.model.save_addition('model_info',model_info)
        # print(str(self.textEdit_cov.toHtml()))
        if hasattr(self, 'covariance_matrix'):
            # self.model.save_addition('covariance_matrix',str(self.textEdit_cov.toHtml()))
            self.model.save_addition('covariance_matrix',self.covariance_matrix)
        else:
            self.model.save_addition('covariance_matrix',pd.DataFrame(np.identity(10)))
        if hasattr(self, 'sensitivity_data'):
            self.model.save_addition('sensitivity',str(self.sensitivity_data))
            # print(str(self.sensitivity_data))
        else:
            self.model.save_addition('sensitivity',str([]))
            # print(pars[i],str(values[i]))
    
    def load_addition(self):
            funcs=\
                [self.widget_solver.par.param('Diff.Ev.').param('k_m').setValue,
                self.widget_solver.par.param('Diff.Ev.').param('k_r').setValue,
                self.widget_solver.par.param('Diff.Ev.').param('Method').setValue,
                self.widget_solver.par.param('FOM').param('Figure of merit').setValue,
                self.widget_solver.par.param('FOM').param('Auto save, interval').setValue,
                self.widget_solver.par.param('FOM').param('weighting factor').setValue,
                self.widget_solver.par.param('FOM').param('weighting region').setValue,
                self.widget_solver.par.param('Fitting').param('start guess').setValue,
                self.widget_solver.par.param('Fitting').param('Generation size').setValue,
                self.widget_solver.par.param('Fitting').param('Population size').setValue]

            types= [float,float,str,str,int,float,str,bool,int,int]
            pars = ['k_m','k_r','Method','Figure of merit','Auto save, interval','weighting factor','weighting region','start guess','Generation size','Population size']
            value = None
            for i in range(len(pars)):
                type_ = types[i]
                if type_ == float:
                    try:
                        value = np.round(float(self.model.load_addition(pars[i])),2)
                    except:
                        pass
                elif type_==str:
                    try:
                        value = self.model.load_addition(pars[i]).decode("utf-8")
                    except:
                        pass
                elif type_==bool:
                    try:
                        value = (self.model.load_addition(pars[i]).decode("ASCII")=="True")
                    except:
                        pass
                else:
                    try:
                        value = type_(self.model.load_addition(pars[i]))
                    except:
                        pass
                if value!=None:
                    funcs[i](value)
            model_info = ''
            sensitivity_data = []
            covariance_matrix = pd.DataFrame(np.identity(3))
            try:
                model_info = self.model.load_addition('model_info').decode('utf-8')
            except:
                pass
            try:
                sensitivity_data = eval(self.model.load_addition('sensitivity').decode('utf-8'))
            except:
                pass
            try:
                covariance_matrix = self.model.load_addition('covariance_matrix', load_type = 'object')
            except:
                pass
            if hasattr(self,'textEdit_note'):
                self.textEdit_note.setPlainText(model_info)
            # self.textEdit_cov.setHtml(covariant_matrix)
            self.sensitivity_data = sensitivity_data
            self.covariance_matrix = covariance_matrix
            self.textEdit_cov.setHtml(covariance_matrix.style.background_gradient(cmap='coolwarm').set_precision(3).render())            
            self.plot_bar_chart(sensitivity_data)

    def simulate_model(self, compile = True):
        """
        simulate the model
        script will be updated and compiled to make name spaces in script_module
        """
        self.update_data_check_attr()
        self.update_plot_dimension()
        self.update_par_upon_change()
        self.model.script = (self.plainTextEdit_script.toPlainText())
        self.widget_solver.update_parameter_in_solver(self)
        self.tableWidget_pars.setShowGrid(True)
        try:
            self.model.simulate(compile = compile)
            self.update_structure_view(compile = compile)
            try:
                self.calc_f_ideal()
            except:
                # self.calc_f_ideal()
                pass
            self.label_2.setText('FOM {}:{}'.format(self.model.fom_func.__name__,self.model.fom))
            self.update_plot_data_view_upon_simulation()
            self.update_electron_density_profile()
            if hasattr(self.model.script_module,'model_type'):
                if self.model.script_module.model_type=='ctr':
                    self.init_structure_view()
                else:
                    pass
            else:
                self.init_structure_view()
            self.statusbar.clearMessage()
            self.update_combo_box_list_par_set()
            self.textBrowser_error_msg.clear()
            self.spinBox_domain.setMaximum(len(self.model.script_module.sample.domain)-1)
            self.statusbar.showMessage("Model is simulated successfully!")
        except model.ModelError as e:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to simulate model!')
            logging.getLogger().exception('Fatal error encountered during model simulation!')
            self.tabWidget_data.setCurrentIndex(4)
            _ = QMessageBox.question(self, 'Runtime error message', str(e), QMessageBox.Ok)

    #execution when you move the slide bar to change only one parameter
    def simulate_model_light(self):
        try:
            self.model.simulate()
            self.label_2.setText('FOM {}:{}'.format(self.model.fom_func.__name__,self.model.fom))
            self.update_plot_data_view_upon_simulation()
            self.update_structure_view(compile = False)
            self.update_electron_density_profile()
            self.statusbar.showMessage("Model is simulated now!")
        except model.ModelError as e:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to simulate model!')
            logging.getLogger().exception('Fatal error encountered during model simulation!')
            self.tabWidget_data.setCurrentIndex(4)
            _ = QMessageBox.question(self, 'Runtime error message', str(e), QMessageBox.Ok)

    def play_with_one_par(self):
        selected_rows = self.tableWidget_pars.selectionModel().selectedRows()
        if len(selected_rows)>0:
            #only get the first selected item
            par_set = self.model.parameters.data[selected_rows[0].row()]
            par_min, par_max = par_set[-4], par_set[-3]
            value = (par_max - par_min)*self.horizontalSlider_par.value()/100 + par_min
            self.model.parameters.set_value(selected_rows[0].row(), 1, value)
            self.lineEdit_scan_par.setText('{}:{}'.format(par_set[0],value))
            self.simulate_model_light()
        else:
            print('Doing nothing!')
            pass

    def scan_one_par(self):
        selected_rows = self.tableWidget_pars.selectionModel().selectedRows()
        if len(selected_rows)>0:
            par_set = self.model.parameters.data[selected_rows[0].row()]
            par_min, par_max = par_set[-4], par_set[-3]
            steps = int(self.spinBox_steps.value())
            for i in range(steps+1):
                value = (par_max - par_min)/steps*i + par_min
                self.model.parameters.set_value(selected_rows[0].row(), 1, value)
                self.horizontalSlider_par.setValue(int(i/steps*100))
                self.lineEdit_scan_par.setText('{}:{}'.format(par_set[0],value))
                self.simulate_model_light()

    def rock_one_par(self, sign):
        selected_rows = self.tableWidget_pars.selectionModel().selectedRows()
        if len(selected_rows)>0:
            par_set = self.model.parameters.data[selected_rows[0].row()]
            par_min, par_max = par_set[-4], par_set[-3]
            steps = int(self.spinBox_steps.value())
            old_value = self.model.parameters.get_value(selected_rows[0].row(), 1)
            new_value = (par_max - par_min)/steps*sign + old_value
            if (new_value>par_max) or (new_value<par_min):
                self.model.parameters.set_value(selected_rows[0].row(), 1, new_value - 2*sign*(par_max - par_min)/steps)
                self.simulate_model_light()
                return -sign
            else:
                self.model.parameters.set_value(selected_rows[0].row(), 1, new_value)
                self.simulate_model_light()
                return sign

    def update_structure_during_scan_par(self):
        self.simulate_model_light()

    def start_scan_par_thread(self):
        selected_rows = self.tableWidget_pars.selectionModel().selectedRows()
        if len(selected_rows)>0:
            self.scan_par.row = selected_rows[0].row()
            self.scan_par.steps = int(self.spinBox_steps.value())
            self.scan_par_thread.start()
            self.timer_scan_par.start(1000)
        else:
            pass

    def stop_scan_par_thread(self):
        self.scan_par.stop()
        self.scan_par_thread.terminate()
        self.timer_scan_par.stop()

    def run_model(self):
        """start the model fit looping"""
        #button will be clicked every 2 second to update figures
        try:
            # self.stop_model()
            self.simulate_model()
            self.statusbar.showMessage("Initializing model running ...")
            self.timer_update_structure.start(2000)
            self.widget_solver.update_parameter_in_solver(self)
            self.fit_thread.start()
        except:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to launch a model fit!')
            logging.getLogger().exception('Fatal error encountered during init model fitting!')
            self.tabWidget_data.setCurrentIndex(5)

    def stop_model(self):
        self.run_fit.stop()
        self.fit_thread.terminate()
        self.timer_update_structure.stop()
        self.statusbar.clearMessage()
        self.statusbar.showMessage("Model run is aborted!")
        
    @QtCore.pyqtSlot(str)
    def stop_model_slot(self,message):
        self.stop_model()
        logging.getLogger().exception(message)
        self.tabWidget_data.setCurrentIndex(5)

    def _stop_model(self):
        self.run_batch.stop()
        self.batch_thread.terminate()
        self.timer_update_structure.stop()
        self.statusbar.clearMessage()
        self.statusbar.showMessage("Batch model run is aborted!")

    def run_model_batch(self):
        """start the model fit looping in a batch mode
        To speed up the structure and plots are not to be updated!
        """
        try:
            #self._stop_model()
            self.simulate_model()
            self.statusbar.clearMessage()
            self.statusbar.showMessage("Initializing model running ...")
            self.widget_solver.update_parameter_in_solver_batch(self)
            self.statusbar.clearMessage()
            self.statusbar.showMessage("Parameters in solver are updated!")
            self.batch_thread.start()
            self.timer_update_structure.start(5000)
        except Exception:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to batch run a model!')
            logging.getLogger().exception('Fatal error encountered during batch run a model!')
            self.tabWidget_data.setCurrentIndex(4)

    def stop_model_batch(self):
        #now need to distinguish how the thread is stopped
        #stop after the finishing all generations
        finish_normal = not self.run_batch.solver.optimizer.stop
        self.update_par_upon_change()
        # self.run_batch.solver.model.simulate()#update the error bar info in the model
        #stop the batch run
        self.run_batch.stop()
        self.batch_thread.terminate()
        self.timer_update_structure.stop()
        #save model first before rolling to next one
        self.calculate_error_bars()
        self.save_model()
        self.auto_save_model()
        #if the run is terminated by user then stop here, otherwise continue on
        if not finish_normal:
            return
        self.statusbar.clearMessage()
        self.statusbar.showMessage("Batch model run is aborted to work on next task!")
        if self.run_batch.multiple_files_hooker:
            if self.rolling_to_next_rod_file():
                self.run_model_batch()
            else:
                pass
        else:
            if self.update_fit_setup_for_batch_run():
                self.run_model_batch()
            else:
                pass

    def terminate_model_batch(self):
        self.run_batch.stop()
        self.batch_thread.terminate()
        self.statusbar.clearMessage()
        self.statusbar.showMessage("Batch model run is aborted now!")

    def rolling_to_next_rod_file(self):
        which = self.run_batch.rod_files.index(self.rod_file)
        if which == self.run_batch.rod_files.__len__()-1:
            return False
        else:
            self.open_model_with_path(self.run_batch.rod_files[which+1])
            self.listWidget_rod_files.setCurrentRow(which+1)
            return True

    def update_fit_setup_for_batch_run(self):
        """
        Update the fit parameters and the fit dataset for next batch job!
        
        Returns:
            [bool] -- move to the end of datasets or not?
        """
        first_checked_data_item, first_checked_par_item = None, None
        for i in range(self.tableWidget_data.rowCount()):
            if self.tableWidget_data.cellWidget(i,2).checkState()!=0:
                first_checked_data_item = i
                break
        for i in range(self.tableWidget_pars.rowCount()):
            if self.tableWidget_pars.cellWidget(i,2)!=None:
                if self.tableWidget_pars.cellWidget(i,2).checkState()!=0:
                    first_checked_par_item = i
                    break
        self.use_none_data()
        self.fit_none()
        try:
            [self.tableWidget_pars.cellWidget(i+6+first_checked_par_item,2).setChecked(True) for i in range(5)]
            self.tableWidget_data.cellWidget(1+first_checked_data_item,2).setChecked(True)
            self.update_model_parameter()
            return True
        except:
            return False

    def load_data(self, loader = 'ctr'):
        self._empty_data_pool()
        exec('self.load_data_{}()'.format(loader))

    def append_data(self):
        self.load_data_ctr()

    def _empty_data_pool(self):
        #now empty the data pool
        self.model.data.items = [data.DataSet(name='Data 0')]
        self.model.data._counter = 1

    def load_data_ctr(self):
        """
        load data
        ------------
        if the data is ctr data, then you should stick to the dataformat as follows
        #8 columns in total
        #X, H, K, Y, I, eI, LB, dL
        #for CTR data, X column is L column, Y column all 0
        #for RAXR data, X column is energy column, Y column is L column
        #H, K, columns are holding H, K values
        #I column is holding background-subtraced intensity of ctr signal
        #LB, and dL are two values for roughness calculation
           LB: first Bragg peak L of one rod
           dL: interval L between two adjacent Bragg peak L's
        To get access to these columns:
            X column: data.x
            I column: data.y
            eI column: data.error
            H column: data.extra_data["h"]
            K column: data.extra_data["k"]
            Y column: data.extra_data["Y"]
            LB column: data.extra_data["LB"]
            dL column: data.extra_data["dL"]
        ---------------
        if the data you want to load is not in CTR format, to make successful loading, assure:
            1)your data file has 8 columns
            2)columns are space-separated (or tab-seperated)
            3)you can add comment lines heading with "#"
            4)if your data has <8 columns, then fill the other unused columns with 0
            5)to asscess your data column, you should use the naming rule described above, although
              the real meaning of each column, eg X column, could be arbitrary at your wishes
              For example, you put frequence values to the first column(X column), then to access this
              column, you use data.X

        Data file of 8 columns should be enough to encountpass many different situations.
        """
        self.model.compiled = False
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","csv Files (*.csv);;data Files (*.dat);txt Files (*.txt)", options=options)
        current_data_set_name = [self.tableWidget_data.item(i,0).text() for i in range(self.tableWidget_data.rowCount())]
        if fileName:
            with open(fileName,'r') as f:
                data_loaded = np.loadtxt(f,comments = '#',delimiter=None)
                data_loaded_pd = pd.DataFrame(data_loaded, columns = ['X','h','k','Y','I','eI','LB','dL'])
                data_loaded_pd['h'] = data_loaded_pd['h'].apply(lambda x:int(np.round(x)))
                data_loaded_pd['k'] = data_loaded_pd['k'].apply(lambda x:int(np.round(x)))
                data_loaded_pd.sort_values(by = ['h','k','Y'], inplace = True)
                hk_unique = list(set(zip(list(data_loaded_pd['h']), list(data_loaded_pd['k']), list(data_loaded_pd['Y']))))
                hk_unique.sort()
                h_unique = [each[0] for each in hk_unique]
                k_unique = [each[1] for each in hk_unique]
                Y_unique = [each[2] for each in hk_unique]

                for i in range(len(h_unique)):
                    h_temp, k_temp, Y_temp = h_unique[i], k_unique[i], Y_unique[i]
                    if Y_temp==0:#CTR data
                        name = 'Data-{}{}L'.format(h_temp, k_temp)
                    else:#RAXR data
                        name = 'Data-{}{}_L={}'.format(h_temp, k_temp, Y_temp)
                    tag = sum([int(name in each) for each in current_data_set_name])+1
                    #if name in current_data_set_name:
                    name = name + '_{}'.format(tag)
                    self.model.data.add_new(name = name)
                    sub_data = data_loaded_pd[(data_loaded_pd['h']==h_temp) & (data_loaded_pd['k']==k_temp)& (data_loaded_pd['Y']==Y_temp)]
                    sub_data.sort_values(by='X',inplace =True)
                    self.model.data.items[-1].x = sub_data['X'].to_numpy()
                    self.model.data.items[-1].y = sub_data['I'].to_numpy()
                    self.model.data.items[-1].error = sub_data['eI'].to_numpy()
                    self.model.data.items[-1].x_raw = sub_data['X'].to_numpy()
                    self.model.data.items[-1].y_raw = sub_data['I'].to_numpy()
                    self.model.data.items[-1].error_raw = sub_data['eI'].to_numpy()
                    self.model.data.items[-1].set_extra_data(name = 'h', value = sub_data['h'].to_numpy())
                    self.model.data.items[-1].set_extra_data(name = 'k', value = sub_data['k'].to_numpy())
                    self.model.data.items[-1].set_extra_data(name = 'Y', value = sub_data['Y'].to_numpy())
                    self.model.data.items[-1].set_extra_data(name = 'LB', value = sub_data['LB'].to_numpy())
                    self.model.data.items[-1].set_extra_data(name = 'dL', value = sub_data['dL'].to_numpy())
                    self.model.data.items[-1].mask = np.array([True]*len(self.model.data.items[-1].x))
                    # self.model.data.concatenate_all_ctr_datasets()
        #now remove the empty datasets
        empty_data_index = []
        i=0
        for each in self.model.data.items:
            if len(each.x_raw) == 0:
                empty_data_index.append(i)
            i += 1
        for i in range(len(empty_data_index)):
            self.model.data.delete_item(empty_data_index[i])
            for ii in range(len(empty_data_index)):
                if empty_data_index[ii]>empty_data_index[i]:
                    empty_data_index[ii] = empty_data_index[ii]-1
                else:
                    pass
        self.model.data_original = copy.deepcopy(self.model.data)
        self.model.data.concatenate_all_ctr_datasets()
        #update the view
        self.update_table_widget_data()
        self.update_combo_box_dataset()
        self.update_plot_dimension()
        self.update_plot_data_view()

    def delete_data(self):
        self.model.compiled = False
        # Delete the selected mytable lines
        row_index = [each.row() for each in self.tableWidget_data.selectionModel().selectedRows()]
        row_index = sorted(row_index, reverse=True)
        for each in row_index:
            self.model.data.delete_item(each)
            self.model.data_original.delete_item(each)
        self.update_table_widget_data()
        self.update_combo_box_dataset()
        self.update_plot_dimension()
        self.update_plot_data_view()

    def update_table_widget_data(self):
        self.tableWidget_data.clear()
        self.tableWidget_data.setRowCount(len(self.model.data))
        self.tableWidget_data.setColumnCount(5)
        self.tableWidget_data.setHorizontalHeaderLabels(['DataID','logY','Use','Errors','fmt'])
        for i in range(len(self.model.data)):
            current_data = self.model.data[i]
            name = current_data.name
            for j in range(5):
                if j == 0:
                    qtablewidget = QTableWidgetItem(name)
                    self.tableWidget_data.setItem(i,j,qtablewidget)
                elif j == 4:
                    qtablewidget = QTableWidgetItem('sym:6bw;l:3r')
                    self.tableWidget_data.setItem(i,j,qtablewidget)
                else:
                    #note j=1 to j=3 corresponds to data.show, data.use, data.use_error
                    #data.show is not used for judging showing or not(all datasets are shown)
                    #It is instead used to specify the scale of Y(log or not)
                    check = getattr(current_data, ['show', 'use', 'use_error'][j-1])
                    check_box = QCheckBox()
                    check_box.setChecked(check)
                    #check_box.stateChanged.connect(self.update_plot_data_view)
                    self.tableWidget_data.setCellWidget(i,j,check_box)
        
        # self.tableWidget_data.resizeColumnsToContents()
        # self.tableWidget_data.resizeRowsToContents()

    def use_all_data(self):
        """fit all datasets
        """
        num_rows_table = self.tableWidget_data.rowCount()
        for i in range(num_rows_table):
            self.tableWidget_data.cellWidget(i,2).setChecked(True)
        self.simulate_model()

    def use_none_data(self):
        """fit none of those datasets
        """
        num_rows_table = self.tableWidget_data.rowCount()
        for i in range(num_rows_table):
            self.tableWidget_data.cellWidget(i,2).setChecked(False)
        self.simulate_model()

    def use_selected_data(self):
        """fit those that have been selected
        """
        selected_row_index = [each.row() for each in self.tableWidget_data.selectionModel().selectedRows()]
        num_rows_table = self.tableWidget_data.rowCount()
        for i in range(num_rows_table):
            if i in selected_row_index:
                self.tableWidget_data.cellWidget(i,2).setChecked(True)
            else:
                self.tableWidget_data.cellWidget(i,2).setChecked(False)
        self.simulate_model()

    def invert_use_data(self):
        """invert the selection of to-be-fit datasets
        """
        num_rows_table = self.tableWidget_data.rowCount()
        for i in range(num_rows_table):
            checkstate = self.tableWidget_data.cellWidget(i,2).checkState()
            if checkstate == 0:
                self.tableWidget_data.cellWidget(i,2).setChecked(True)
            else:
                self.tableWidget_data.cellWidget(i,2).setChecked(False)
        self.simulate_model()

    def update_combo_box_dataset(self):
        new_items = [each.name for each in self.model.data]
        self.comboBox_dataset.clear()
        self.comboBox_dataset.addItems(new_items)
        self.comboBox_dataset2.clear()
        self.comboBox_dataset2.addItems(new_items)

    #used in q correction
    def return_L_I(self):
        dataset_name = self.comboBox_dataset2.currentText()
        dataset = None
        for each in self.model.data_original:
            if each.name == dataset_name:
                dataset = each
                break
            else:
                pass
        return dataset.x, dataset.y

    def update_data_view(self):
        """update the data view widget to show data values as table"""
        dataset_name = self.comboBox_dataset.currentText()
        dataset = None
        for each in self.model.data_original:
            if each.name == dataset_name:
                dataset = each
                break
            else:
                pass
        column_labels_main = ['x','y','error','mask']
        extra_labels = ['h', 'k', 'dL', 'LB']
        all_labels = ['x','y','error','h','k','dL','LB','mask']
        self.tableWidget_data_view.setRowCount(len(dataset.x))
        self.tableWidget_data_view.setColumnCount(len(all_labels))
        self.tableWidget_data_view.setHorizontalHeaderLabels(all_labels)
        for i in range(len(dataset.x)):
            for j in range(len(all_labels)):
                if all_labels[j] in column_labels_main:
                    item_ = getattr(dataset,all_labels[j])[i]
                    if all_labels[j] == 'mask':
                        qtablewidget = QTableWidgetItem(str(item_))
                    else:
                        qtablewidget = QTableWidgetItem(str(round(item_,4)))
                elif all_labels[j] in extra_labels:
                    qtablewidget = QTableWidgetItem(str(dataset.get_extra_data(all_labels[j])[i]))
                else:
                    qtablewidget = QTableWidgetItem('True')
                self.tableWidget_data_view.setItem(i,j,qtablewidget)

    def update_mask_info_in_data(self):
        """if the mask value is False, the associated data point wont be shown and wont be fitted as well"""
        dataset_name = self.comboBox_dataset.currentText()
        dataset = None
        for each in self.model.data_original:
            if each.name == dataset_name:
                dataset = each
                break
            else:
                pass
        for i in range(len(dataset.x)):
            dataset.mask[i] = (self.tableWidget_data_view.item(i,7).text() == 'True')
        self.model.data = copy.deepcopy(self.model.data_original)
        [each.apply_mask() for each in self.model.data]
        #updae the data infomation
        self.model.data.concatenate_all_ctr_datasets()
        self.simulate_model()

    def init_mask_info_in_data_upon_loading_model(self):
        """apply mask values to each dataset"""
        self.model.data = copy.deepcopy(self.model.data_original)
        [each.apply_mask() for each in self.model.data]
        self.simulate_model()

    def init_structure_view(self):
        try:
            domain_tag = int(self.spinBox_domain.text())
        except:
            domain_tag = 0
        size_domain = len(self.model.script_module.sample.domain)
        if size_domain<(1+domain_tag):
            domain_tag = size_domain -1
        else:
            pass
        # self.widget_edp.items = []
        # self.widget_msv_top.items = []
        self.widget_edp.abc = [self.model.script_module.sample.unit_cell.a,self.model.script_module.sample.unit_cell.b,self.model.script_module.sample.unit_cell.c]
        self.widget_edp.T = self.model.script_module.sample.unit_cell.lattice.RealTM
        self.widget_edp.T_INV = self.model.script_module.sample.unit_cell.lattice.RealTMInv
        self.widget_edp.super_cell_size = eval(self.lineEdit_super_cell.text())
        self.widget_edp.show_bond_length = self.checkBox_label.isChecked()
        # self.widget_msv_top.abc = self.widget_edp.abc
        xyz = self.model.script_module.sample.extract_xyz_top(domain_tag, num_of_atomic_layers = self.spinBox_layers.value(), use_sym = self.checkBox_symmetry.isChecked(),size = eval(self.lineEdit_super_cell.text()))
        self.widget_edp.show_structure(xyz)
        try:
            azimuth = self.widget_edp.opts['azimuth']
            elevation = self.widget_edp.opts['elevation']
        except:
            azimuth, elevation = 0, 0
        self.update_camera_position(widget_name = 'widget_edp', angle_type="azimuth", angle=azimuth)
        self.update_camera_position(widget_name = 'widget_edp', angle_type = 'elevation', angle = elevation)
        self.update_electron_density_profile()

        # xyz,_ = self.model.script_module.sample.extract_xyz_top(domain_tag)
        # self.widget_msv_top.show_structure(xyz)
        # self.update_camera_position(widget_name = 'widget_msv_top', angle_type="azimuth", angle=0)
        # self.update_camera_position(widget_name = 'widget_msv_top', angle_type = 'elevation', angle = 90)
        """
        try:
            xyz,_ = self.model.script_module.sample.extract_xyz_top(domain_tag)
            self.widget_msv_top.show_structure(xyz)
            self.update_camera_position(widget_name = 'widget_msv_top', angle_type="azimuth", angle=0)
            self.update_camera_position(widget_name = 'widget_msv_top', angle_type = 'elevation', angle = 90)
        except:
            pass
        """

    def update_structure_view(self, compile = True):
        if hasattr(self.model.script_module,"model_type"):
            if getattr(self.model.script_module,"model_type")=="ctr":
                pass
            else:
                return
        else:
            pass
        try:
            if self.spinBox_domain.text()=="":
                domain_tag = 0
            else:
                domain_tag = int(self.spinBox_domain.text())
            size_domain = len(self.model.script_module.sample.domain)
            if size_domain<(1+domain_tag):
                domain_tag = size_domain -1
            else:
                pass        
            xyz = self.model.script_module.sample.extract_xyz_top(domain_tag, num_of_atomic_layers = self.spinBox_layers.value(),use_sym = self.checkBox_symmetry.isChecked(),size = eval(self.lineEdit_super_cell.text()))
            if self.run_fit.running or (not compile): 
                self.widget_edp.update_structure(xyz)
            else:
                self.widget_edp.clear()
                #self.widget_edp.items = []
                self.widget_edp.abc = [self.model.script_module.sample.unit_cell.a,self.model.script_module.sample.unit_cell.b,self.model.script_module.sample.unit_cell.c]
                self.widget_edp.T = self.model.script_module.sample.unit_cell.lattice.RealTM
                self.widget_edp.T_INV = self.model.script_module.sample.unit_cell.lattice.RealTMInv
                self.widget_edp.super_cell_size = eval(self.lineEdit_super_cell.text())
                self.widget_edp.show_bond_length = self.checkBox_label.isChecked()
                self.widget_edp.show_structure(xyz)
            #let us also update the eden profile
            self.update_electron_density_profile()

            """
            try:
                xyz, _ = self.model.script_module.sample.extract_xyz_top(domain_tag)
                self.widget_msv_top.update_structure(xyz)
            except:
                pass
            """
        except Exception as e:
            outp = StringIO()
            traceback.print_exc(200, outp)
            val = outp.getvalue()
            outp.close()
            _ = QMessageBox.question(self, "",'Runtime error message:\n{}'.format(str(val)), QMessageBox.Ok)

    def save_structure_file(self):
        domain_tag, done = QInputDialog.getInt(self, 'Domain tag', 'Enter the domain index for the structure you want to save eg 0:')
        if not done:
            domain_tag = 0
        path, _ = QFileDialog.getSaveFileName(self, "Save file", "", "xyz file (*.xyz)")
        self.model.script_module.sample.make_xyz_file(which_domain = int(domain_tag), save_file = path)
        self.statusbar.clearMessage()
        self.statusbar.showMessage('The data file is saved at {}'.format(path))

    def save_data(self):
        def _make_data():
            """[append missing points near Bragg peaks, the values for data column at these points will be set to nan, while the values for model column will be calculated]
            """
            extended_data = {}
            keys = ['potential','L', 'H', 'K', 'I', 'I_model', 'error', 'I_bulk', 'use']
            for each in keys:
                extended_data[each] = []
            for data in self.model.data:
                L = data.x
                H = data.extra_data['h']
                K = data.extra_data['k']
                dL = data.extra_data['dL']
                LB = data.extra_data['LB']
                I = data.y
                #I_model = data.y_sim
                error = data.error
                Bragg_L = LB[0] + np.array(range(-2,10))*dL[0]
                Bragg_L = [each for each in Bragg_L if L.max()>each>L.min()]
                Bragg_index = []
                for each_bragg_L in Bragg_L:
                    ix = np.argpartition(abs(L - each_bragg_L),1)
                    left, right = None, None
                    ix_left, ix_right = None, None
                    if L[ix[0]]>each_bragg_L:
                        right = L[ix[0]]
                        ix_right = ix[0]
                        left = L[ix[0]-1]
                        ix_left = ix_right -1
                    else:
                        left = L[ix[0]]
                        ix_left = ix[0]
                        right = L[ix[0]+1]
                        ix_right = ix[0]+1
                    Bragg_index.append([ix_left+num_points_near_Bragg_peak, ix_left+num_points_near_Bragg_peak+1])
                    appended_Ls = list(np.linspace(left, each_bragg_L-0.02, num_points_near_Bragg_peak, endpoint = True))+ list(np.linspace(right, each_bragg_L+0.02, num_points_near_Bragg_peak, endpoint = True))[::-1]
                    appended_Hs = [H[0]]*len(appended_Ls)
                    appended_Ks = [K[0]]*len(appended_Ls)
                    appended_dL = [dL[ix_right]]*len(appended_Ls)
                    appended_LB = [LB[ix_right]]*len(appended_Ls)
                    L = np.concatenate((L[:ix_right],appended_Ls,L[ix_right:]))
                    H = np.concatenate((H[:ix_right],[H[0]]*len(appended_Ls),H[ix_right:]))
                    K = np.concatenate((K[:ix_right],[K[0]]*len(appended_Ls),K[ix_right:]))
                    dL = np.concatenate((dL[:ix_right],[dL[ix_right]]*len(appended_Ls),dL[ix_right:]))
                    LB = np.concatenate((LB[:ix_right],[LB[ix_right]]*len(appended_Ls),LB[ix_right:]))
                    I = np.concatenate((I[:ix_right],[np.nan]*len(appended_Ls),I[ix_right:]))
                    #I_model = np.concatenate((I_model[:ix_right],[np.nan]*len(appended_Ls),I_model[ix_right:]))
                    error = np.concatenate((error[:ix_right],[np.nan]*len(appended_Ls),error[ix_right:]))
                beta = self.model.script_module.rgh.beta
                rough = (1-beta)/((1-beta)**2 + 4*beta*np.sin(np.pi*(L-LB)/dL)**2)**0.5
                f = rough*self.model.script_module.sample.calc_f_all(H, K, L)
                f_ideal = self.model.script_module.sample.calc_f_ideal(H, K, L)
                extra_scale_factor = 'scale_factor_{}{}L'.format(int(round(H[0],0)),int(round(K[0],0)))
                if hasattr(self.model.script_module.rgh,extra_scale_factor):
                    rod_factor = getattr(self.model.script_module.rgh, extra_scale_factor)
                else:
                    if int(round(H[0],0))==0 and int(round(K[0],0))==0:#specular rod
                        if hasattr(self.model.script_module.rgh,'scale_specular_rod'):
                            rod_factor = getattr(self.model.script_module.rgh,'scale_specular_rod')
                        else:
                            rod_factor = 1
                    else:#nonspecular rod
                        if hasattr(self.model.script_module.rgh,'scale_nonspecular_rod'):
                            rod_factor = getattr(self.model.script_module.rgh,'scale_nonspecular_rod')
                        else:
                            rod_factor = 1
                I_model = list(abs(f*f)*self.model.script_module.rgh.scale_nonspecular_rods*rod_factor)
                I_bulk = list(abs(f_ideal*f_ideal)*self.model.script_module.rgh.scale_nonspecular_rods*rod_factor)
                E = [potential]*len(L)
                use = [True]*len(L)
                for each in keys:
                    if each=='potential':
                        new = locals()['E']
                    else:
                        new = locals()[each]
                    extended_data[each] = list(extended_data[each]) + list(new)
            return extended_data

        num_points_near_Bragg_peak = 4
        potential, done = QInputDialog.getDouble(self, 'Potential_info', 'Enter the potential for this dataset (in V):')
        if not done:
            potential = None
        path, _ = QFileDialog.getSaveFileName(self, "Save file", "", "data file (*.*)")
        if path!="":
            export_data = _make_data()
            df_export_data = pd.DataFrame(export_data)
            self._append_df_to_excel(filename = [path+'.xlsx',path][int(path.endswith('.xlsx'))], df = df_export_data, sheet_name = 'Sheet1', startrow = None, truncate_sheet=False, columns = ['potential','L', 'H', 'K', 'I', 'I_model', 'error', 'I_bulk', 'use'])
            self.save_data_original(path=path)
            '''
            #also save loadable csv file
            #df_export_data.to_csv([path+'.csv',path][int(path.endswith('.csv'))],sep="\t",columns=['L','H','K','Y','I','error','LB','dL'],\
                                 #index=False, header=['#L','H','K','Y','I','error','LB','dL'])
            '''

    def _append_df_to_excel(self, filename, df, sheet_name='Sheet1', startrow=None,
                        truncate_sheet=False, 
                        **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
        filename : File path or existing ExcelWriter
                    (Example: '/path/to/file.xlsx')
        df : dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
        startrow : upper left cell row to dump data frame.
                    Per default (startrow=None) calculate the last row
                    in the existing DF and write to the next row...
        truncate_sheet : truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
        to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')
        header = None

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
            header = False
        except FileNotFoundError:
            # file does not exist yet, we will create it
            header = True
            #pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, header = header, **to_excel_kwargs)

        # save the workbook
        writer.save()

    #save data plus best fit profile
    def save_data_original(self, path=""):
        '''
        potential, done = QInputDialog.getDouble(self, 'Potential_info', 'Enter the potential for this dataset (in V):')
        if not done:
            potential = None
        path, _ = QFileDialog.getSaveFileName(self, "Save file", "", "data file (*.*)")
        '''
        if path!="":
            keys_attri = ['x','y','y_sim','error']
            keys_extra = ['h','k','Y','dL','LB']
            lib_map = {'x': 'L', 'y':'I','y_sim':'I_model','error':'error','h':'H','k':'K','Y':'Y','dL':'dL','LB':'LB'}
            export_data = {}
            for key in ['x','h','k','y','y_sim','error','Y','dL','LB']:
                export_data[lib_map[key]] = []
            export_data['use'] = []
            #export_data['I_bulk'] = []
            #export_data['potential'] = []
            for each in self.model.data:
                if each.use:
                    for key in ['x','h','k','y','y_sim','error','Y','dL','LB']:
                        if key in keys_attri:
                            export_data[lib_map[key]] = np.append(export_data[lib_map[key]], getattr(each,key))
                        elif key in keys_extra:
                            export_data[lib_map[key]] = np.append(export_data[lib_map[key]], each.extra_data[key])
                    export_data['use'] = np.append(export_data['use'],[True]*len(each.x))
                else:
                    for key in ['x','h','k','y','y_sim','error','Y','dL','LB']:
                        if key in keys_attri:
                            if key=='y_sim':
                                export_data[lib_map[key]] = np.append(export_data[lib_map[key]], [0]*len(getattr(each,'x')))
                            else:
                                export_data[lib_map[key]] = np.append(export_data[lib_map[key]], getattr(each,key))
                        elif key in keys_extra:
                            export_data[lib_map[key]] = np.append(export_data[lib_map[key]], each.extra_data[key])
                    export_data['use'] = np.append(export_data['use'],[False]*len(each.x))
                '''
                export_data['potential'] = np.append(export_data['potential'],[float(potential)]*len(each.x))
                beta = self.model.script_module.rgh.beta
                #rough = (1-beta)/((1-beta)**2 + 4*beta*np.sin(np.pi*(each.x-each.extra_data['LB'])/each.extra_data['dL'])**2)**0.5
                scale_factor = [self.model.script_module.rgh.scale_nonspecular_rods,self.model.script_module.rgh.scale_specular_rod][int("00L" in each.name)]
                h_, k_ = int(round(each.extra_data['h'][0],0)),int(round(each.extra_data['k'][0],0))
                extra_scale_factor = 'scale_factor_{}{}L'.format(h_,k_)
                if hasattr(self.model.script_module.rgh,extra_scale_factor):
                    rod_factor = getattr(self.model.script_module.rgh, extra_scale_factor)
                else:
                    rod_factor = 1
                rough = 1
                export_data['I_bulk'] = np.append(export_data['I_bulk'],rough**2*np.array(self.model.script_module.sample.calc_f_ideal(each.extra_data['h'], each.extra_data['k'], each.x)**2*scale_factor*rod_factor))
                '''
            '''
            writer_temp = pd.ExcelWriter([path+'.xlsx',path][int(path.endswith('.xlsx'))])
            df_export_data.to_excel(writer_temp, columns =['potential']+[lib_map[each_] for each_ in ['x','h','k','y','y_sim','error']]+['I_bulk','use'])
            writer_temp.save()
            writer_temp.close()
            '''
            #also save loadable csv file
            df_export_data = pd.DataFrame(export_data)
            df_export_data.to_csv([path+'.csv',path][int(path.endswith('.csv'))],sep="\t",columns=['L','H','K','Y','I','error','LB','dL'],\
                                 index=False, header=['#L','H','K','Y','I','error','LB','dL'])

    #not implemented!
    def change_plot_style(self):
        if self.background_color == 'w':
            self.widget_data.getViewBox().setBackgroundColor('k')
            self.widget_edp.getViewBox().setBackgroundColor('k')
            # self.widget_msv_top.getViewBox().setBackgroundColor('k')
            self.background_color = 'k'
        else:
            self.widget_data.getViewBox().setBackgroundColor('w')
            self.widget_edp.getViewBox().setBackgroundColor('w')
            # self.widget_msv_top.getViewBox().setBackgroundColor('w')
            self.background_color = 'w'

    def load_script(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","script Files (*.py);;text Files (*.txt)", options=options)
        if fileName:
            with open(fileName,'r') as f:
                self.plainTextEdit_script.setPlainText(f.read())
        self.model.script = (self.plainTextEdit_script.toPlainText())

    def update_script_upon_load(self):
        self.plainTextEdit_script.setPlainText(self.model.script)

    def save_script(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save script file", "", "script file (*.py)")
        with open(path,'w') as f:
            f.write(self.model.script)

    def modify_script(self):
        """
        Modify script based on the specified sorbates and total domain numbers
        To use this function, your script file should be standadized to have 
        tags to specifpy the code block position where you define the sorbates.
        This func is customized to modify script_model_standard.py.
        """
        assert self.model.script!="","No script to work on, please load script first!"
        domain_num = int(self.lineEdit_domain_number.text().rstrip())
        motif_chain = self.lineEdit_sorbate_motif.text().strip().rsplit(",")

        assert domain_num == len(motif_chain), "Number of domain not match with the motif number. Fix it first!!"
        lines = script_block_modifier(self.model.script.rsplit("\n"), 'slabnumber',["num_surface_slabs"],[domain_num])

        els_sorbate = []
        anchor_index_list = []
        flat_down_index = []
        xyzu_oc_m = []
        structure = []
        for each in motif_chain:
            each = each.strip()
            properties_temp = getattr(sorbate_tool,each)
            for each_key in properties_temp:
                if each_key == "els_sorbate":
                    els_sorbate.append(properties_temp[each_key])
                elif each_key == "anchor_index_list":
                    anchor_index_list.append(properties_temp[each_key])
                elif each_key == "flat_down_index":
                    flat_down_index.append(properties_temp[each_key])
                elif each_key == "structure":
                    structure.append("#"+each+properties_temp[each_key])
        xyzu_oc_m = [[0.5, 0.5, 1.5, 0.1, 1, 1]]*len(els_sorbate)
        tag_list = ['els_sorbate', 'anchor_index_list', 'flat_down_index', 'xyzu_oc_m']
        tag_value_list = [els_sorbate, anchor_index_list, flat_down_index, xyzu_oc_m]
        lines = script_block_modifier(lines, 'sorbateproperties',tag_list, tag_value_list)
        left_, right_ = locate_tag(lines,'sorbatestructure')
        del(lines[left_:right_])
        if structure[-1][-1] == "\n":
            structure[-1] = structure[-1][0:-1]
        lines.insert(left_,"\n".join(structure))

        self.model.script = '\n'.join(lines)
        self.plainTextEdit_script.setPlainText(self.model.script)

    def remove_selected_rows(self):
        # Delete the selected mytable lines
        self._deleteRows(self.tableWidget_pars.selectionModel().selectedRows(), self.tableWidget_pars)
        self.update_model_parameter()

    # DeleteRows function
    def _deleteRows(self, rows, table):
            # Get all row index
            indexes = []
            for row in rows:
                indexes.append(row.row())

            # Reverse sort rows indexes
            indexes = sorted(indexes, reverse=True)

            # Delete rows
            for rowidx in indexes:
                table.removeRow(rowidx)

    def append_one_row(self):
        rows = self.tableWidget_pars.selectionModel().selectedRows()
        if len(rows) == 0:
            row_index = self.tableWidget_pars.rowCount()
        else:
            row_index = rows[-1].row()
        self.tableWidget_pars.insertRow(row_index+1)
        for i in range(7):
            if i==2:
                check_box = QCheckBox()
                check_box.setChecked(False)
                self.tableWidget_pars.setCellWidget(row_index+1,2,check_box)
            else:
                qtablewidget = QTableWidgetItem('')
                if i == 0:
                    qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                elif i == 1:
                    qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                self.tableWidget_pars.setItem(row_index+1,i,qtablewidget)
        self.update_model_parameter()

    def append_one_row_at_the_end(self):
        row_index = self.tableWidget_pars.rowCount()
        self.tableWidget_pars.insertRow(row_index)
        for i in range(7):
            if i==2:
                check_box = QCheckBox()
                check_box.setChecked(False)
                self.tableWidget_pars.setCellWidget(row_index,2,check_box)
            else:
                qtablewidget = QTableWidgetItem('')
                if i == 0:
                    qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                elif i == 1:
                    qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                self.tableWidget_pars.setItem(row_index,i,qtablewidget)
        self.update_model_parameter()

    def update_model_parameter(self):
        """After you made changes in the par table, this func is executed to update the par values in model"""
        self.model.parameters.data = []
        vertical_label = []
        label_tag=1
        for i in range(self.tableWidget_pars.rowCount()):
            if self.tableWidget_pars.item(i,0)==None:
                items = ['',0,False,0,0,'-','']
                vertical_label.append('')
            elif self.tableWidget_pars.item(i,0).text()=='':
                items = ['',0,False,0,0,'-','']
                vertical_label.append('')
            else:
                items = [self.tableWidget_pars.item(i,0).text(),float(self.tableWidget_pars.item(i,1).text()),self.tableWidget_pars.cellWidget(i,2).isChecked(),\
                         float(self.tableWidget_pars.item(i,3).text()), float(self.tableWidget_pars.item(i,4).text()), self.tableWidget_pars.item(i,5).text()]
                if self.tableWidget_pars.item(i,6)==None:
                    items.append('')
                else:
                    items.append(self.tableWidget_pars.item(i,6).text())
                self.model.parameters.data.append(items)
                if self.tableWidget_pars.cellWidget(i,2).isChecked():
                    vertical_label.append(str(label_tag))
                    label_tag += 1
                else:
                    vertical_label.append('')
        self.tableWidget_pars.setVerticalHeaderLabels(vertical_label)

    def fit_all(self):
        """fit all fit parameters
        """
        num_rows_table = self.tableWidget_pars.rowCount()
        for i in range(num_rows_table):
            try:
                self.tableWidget_pars.cellWidget(i,2).setChecked(True)
            except:
                pass
        self.update_model_parameter()

    def fit_next_5(self):
        """fit next 5 parameters starting from first selected row
        """
        num_rows_table = 5
        rows = self.tableWidget_pars.selectionModel().selectedRows()
        starting_row = 0
        if len(rows)!=0:
            starting_row = rows[0].row()

        for i in range(num_rows_table):
            try:
                self.tableWidget_pars.cellWidget(i+starting_row,2).setChecked(True)
            except:
                pass
        self.update_model_parameter()

    def fit_none(self):
        """fit none of parameters
        """
        num_rows_table = self.tableWidget_pars.rowCount()
        for i in range(num_rows_table):
            try:
                self.tableWidget_pars.cellWidget(i,2).setChecked(False)
            except:
                pass
        self.update_model_parameter()

    def fit_selected(self):
        """fit selected parameters
        """
        selected_row_index = [each.row() for each in self.tableWidget_pars.selectionModel().selectedRows()]
        num_rows_table = self.tableWidget_pars.rowCount()
        for i in range(num_rows_table):
            if i in selected_row_index:
                try:
                    self.tableWidget_pars.cellWidget(i,2).setChecked(True)
                except:
                    pass
            else:
                try:
                    self.tableWidget_pars.cellWidget(i,2).setChecked(False)
                except:
                    pass
        self.update_model_parameter()

    def invert_fit(self):
        """invert the selection of fit parameters
        """
        num_rows_table = self.tableWidget_pars.rowCount()
        for i in range(num_rows_table):
            try:
                checkstate = self.tableWidget_pars.cellWidget(i,2).checkState()
                if checkstate == 0:
                    self.tableWidget_pars.cellWidget(i,2).setChecked(True)
                else:
                    self.tableWidget_pars.cellWidget(i,2).setChecked(False)
            except:
                pass
        self.update_model_parameter()

    def update_par_upon_load(self):
        """upon loading model, the par table widget content will be updated with this func"""
        vertical_labels = []
        lines = self.model.parameters.data
        how_many_pars = len(lines)
        self.tableWidget_pars.clear()
        self.tableWidget_pars.setRowCount(how_many_pars)
        self.tableWidget_pars.setColumnCount(7)
        self.tableWidget_pars.setHorizontalHeaderLabels(['Parameter','Value','Fit','Min','Max','Error','Link'])
        for i in range(len(lines)):
            items = lines[i]
            #j = 0
            if items[0] == '':
                vertical_labels.append('')
                # j += 1
            else:
                #add items to table view
                if len(vertical_labels)==0:
                    if items[2]:
                        vertical_labels.append('1')
                    else:
                        vertical_labels.append('')
                else:
                    #if vertical_labels[-1] != '':
                    if items[2]:#ture or false
                        if '1' not in vertical_labels:
                            vertical_labels.append('1')
                        else:
                            jj=0
                            while vertical_labels[-1-jj]=='':
                                jj = jj + 1
                            vertical_labels.append('{}'.format(int(vertical_labels[-1-jj])+1))
                    else:
                        vertical_labels.append('')
                for j,item in enumerate(items):
                    if j == 2:
                        check_box = QCheckBox()
                        check_box.setChecked(item==True)
                        self.tableWidget_pars.setCellWidget(i,2,check_box)
                    else:
                        if j == 1:
                            qtablewidget = QTableWidgetItem(str(round(item,10)))
                        else:
                            qtablewidget = QTableWidgetItem(str(item))
                        if j == 0:
                            qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                        elif j == 1:
                            qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                        self.tableWidget_pars.setItem(i,j,qtablewidget)
                    #j += 1
        self.tableWidget_pars.resizeColumnsToContents()
        self.tableWidget_pars.resizeRowsToContents()
        """
        header = self.tableWidget_pars.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
        """
        self.tableWidget_pars.setShowGrid(True)
        self.tableWidget_pars.setVerticalHeaderLabels(vertical_labels)

    def _load_par(self):
        vertical_labels = []
        self.tableWidget_pars.setRowCount(1)
        self.tableWidget_pars.setColumnCount(7)
        self.tableWidget_pars.setHorizontalHeaderLabels(['Parameter','Value','Fit','Min','Max','Error','Link'])
        items = ['par',0,'False',0,0,'-','']
        for i in [0]:
            j = 0
            if items[0] == '':
                self.model.parameters.data.append([items[0],0,False,0, 0,'-',''])
                vertical_labels.append('')
                j += 1
            else:
                #add items to parameter attr
                self.model.parameters.data.append([items[0],float(items[1]),items[2]=='True',float(items[3]), float(items[4]),items[5],items[6]])
                #add items to table view
                if len(vertical_labels)==0:
                    vertical_labels.append('1')
                else:
                    if vertical_labels[-1] != '':
                        vertical_labels.append('{}'.format(int(vertical_labels[-1])+1))
                    else:
                        vertical_labels.append('{}'.format(int(vertical_labels[-2])+1))
                for item in items:
                    if j == 2:
                        check_box = QCheckBox()
                        check_box.setChecked(item=='True')
                        self.tableWidget_pars.setCellWidget(i,2,check_box)
                    else:
                        qtablewidget = QTableWidgetItem(item)
                        # qtablewidget.setTextAlignment(Qt.AlignCenter)
                        if j == 0:
                            qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                        elif j == 1:
                            qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                        self.tableWidget_pars.setItem(i,j,qtablewidget)
                    j += 1
        # self.tableWidget_pars.resizeColumnsToContents()
        # self.tableWidget_pars.resizeRowsToContents()
        self.tableWidget_pars.setShowGrid(True)
        self.tableWidget_pars.setVerticalHeaderLabels(vertical_labels)

    def load_par(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","Table Files (*.tab);;text Files (*.txt)", options=options)
        vertical_labels = []
        if fileName:
            with open(fileName,'r') as f:
                lines = f.readlines()
                lines = [each for each in lines if not each.startswith('#')]
                how_many_pars = len(lines)
                self.tableWidget_pars.setRowCount(how_many_pars)
                self.tableWidget_pars.setColumnCount(7)
                self.tableWidget_pars.setHorizontalHeaderLabels(['Parameter','Value','Fit','Min','Max','Error','Link'])
                for i in range(len(lines)):
                    line = lines[i]
                    items = line.rstrip().rsplit('\t')
                    j = 0
                    if items[0] == '':
                        self.model.parameters.data.append([items[0],0,False,0, 0,'-',''])
                        vertical_labels.append('')
                        j += 1
                    else:
                        #add items to parameter attr
                        if len(items)==6:
                            items.append('')
                        self.model.parameters.data.append([items[0],float(items[1]),items[2]=='True',float(items[3]), float(items[4]),items[5],items[6]])
                        #add items to table view
                        if len(vertical_labels)==0:
                            vertical_labels.append('1')
                        else:
                            if vertical_labels[-1] != '':
                                vertical_labels.append('{}'.format(int(vertical_labels[-1])+1))
                            else:
                                vertical_labels.append('{}'.format(int(vertical_labels[-2])+1))
                        for item in items:
                            if j == 2:
                                check_box = QCheckBox()
                                check_box.setChecked(item=='True')
                                self.tableWidget_pars.setCellWidget(i,2,check_box)
                            else:
                                qtablewidget = QTableWidgetItem(item)
                                # qtablewidget.setTextAlignment(Qt.AlignCenter)
                                if j == 0:
                                    qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                                elif j == 1:
                                    qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                                self.tableWidget_pars.setItem(i,j,qtablewidget)
                            j += 1
        self.tableWidget_pars.resizeColumnsToContents()
        self.tableWidget_pars.resizeRowsToContents()
        """
        header = self.tableWidget_pars.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
        """
        self.tableWidget_pars.setShowGrid(True)
        self.tableWidget_pars.setVerticalHeaderLabels(vertical_labels)

    def update_par_upon_change(self):
        """will be executed before simulation"""
        self.model.parameters.data = []
        for each_row in range(self.tableWidget_pars.rowCount()):
            if self.tableWidget_pars.item(each_row,0)==None:
                items = ['',0,False,0,0,'-','']
            elif self.tableWidget_pars.item(each_row,0).text()=='':
                items = ['',0,False,0,0,'-','']
            else:
                items = [self.tableWidget_pars.item(each_row,0).text()] + [float(self.tableWidget_pars.item(each_row,i).text()) for i in [1,3,4]] + [self.tableWidget_pars.item(each_row,5).text()]
                if self.tableWidget_pars.item(each_row,6)!=None:
                    items.append(self.tableWidget_pars.item(each_row,6).text())
                else:
                    items.append('')
                items.insert(2, self.tableWidget_pars.cellWidget(each_row,2).isChecked())
            self.model.parameters.data.append(items)

    @QtCore.pyqtSlot(str,object,bool)
    def update_par_during_fit(self,string,model,save_tag):
        """slot func to update par table widgets during fit"""
        for i in range(len(model.parameters.data)):
            if model.parameters.data[i][0] !='':
                item_temp = self.tableWidget_pars.item(i,1)
                item_temp.setText(str(round(model.parameters.data[i][1],8)))
        self.tableWidget_pars.resizeColumnsToContents()
        self.tableWidget_pars.resizeRowsToContents()
        self.tableWidget_pars.setShowGrid(False)

    @QtCore.pyqtSlot(str,object,bool)
    def update_status(self,string,model,save_tag):
        """slot func to update status info displaying fit status"""
        self.statusbar.clearMessage()
        self.statusbar.showMessage(string)
        self.label_2.setText('FOM {}:{}'.format(self.model.fom_func.__name__,round(self.run_fit.solver.optimizer.best_fom,5)))
        if save_tag:
            self.auto_save_model()

    @QtCore.pyqtSlot(str,object,bool)
    def update_status_batch(self,string,model,save_tag):
        """slot func to update status info displaying fit status"""
        self.statusbar.clearMessage()
        self.statusbar.showMessage(string)
        self.label_2.setText('FOM {}:{}'.format(self.model.fom_func.__name__,round(self.run_batch.solver.optimizer.best_fom,5)))
        if save_tag:
            self.auto_save_model()

    def save_par(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save tab file", "", "table file (*.*)")
        with open(path,'w') as f:
            f.write(self.model.parameters.get_ascii_output())

if __name__ == "__main__":
    QApplication.setStyle("windows")
    app = QApplication(sys.argv)
    #get dpi info: dots per inch
    screen = app.screens()[0]
    dpi = screen.physicalDotsPerInch()
    myWin = MyMainWindow()
    myWin.setWindowIcon(QtGui.QIcon('DAFY.png'))
    hightlight = syntax_pars.PythonHighlighter(myWin.plainTextEdit_script.document())
    myWin.plainTextEdit_script.show()
    myWin.plainTextEdit_script.setPlainText(myWin.plainTextEdit_script.toPlainText())
    #which style would you like to use?
    now = datetime.datetime.now()
    today8am = now.replace(hour=8, minute=0, second=0, microsecond=0)
    today5pm = now.replace(hour=17, minute=0, second=0, microsecond=0)
    if not (today5pm>now>today8am):
        app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    myWin.show()
    sys.exit(app.exec_())
